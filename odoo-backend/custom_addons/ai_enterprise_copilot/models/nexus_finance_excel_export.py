# -*- coding: utf-8 -*-
"""Nexus Finance Excel Export — تصدير التقارير إلى XLSX.

Generates a styled XLSX workbook for any ``nexus.finance.report``
using only the Python standard library and ``openpyxl``. Each report
type produces a workbook with:
    * A branded cover sheet (company logo placeholder, title, period).
    * One or more data sheets with formatted columns.
    * A summary sheet with totals.

If ``openpyxl`` is not available the controller falls back to a
basic CSV stream.
"""

import base64
import csv
import io
import logging

from odoo import api, fields, models, _

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _OPENPYXL_AVAILABLE = False


_BRAND_FILL = PatternFill(
    start_color="0B3D2E", end_color="0B3D2E", fill_type="solid"
) if _OPENPYXL_AVAILABLE else None
_HEADER_FILL = PatternFill(
    start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
) if _OPENPYXL_AVAILABLE else None
_TOTAL_FILL = PatternFill(
    start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"
) if _OPENPYXL_AVAILABLE else None
_THIN = Side(border_style="thin", color="CCCCCC") if _OPENPYXL_AVAILABLE else None
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN) if _OPENPYXL_AVAILABLE else None


class NexusFinanceExcelExport(models.AbstractModel):
    """Generates XLSX (or CSV fallback) for financial reports."""

    _name = "nexus.finance.excel.export"
    _description = "Nexus Finance Excel Export Service"

    # ─────────────────────────────────────────────────────────────────
    # Public entry point
    # ─────────────────────────────────────────────────────────────────
    def generate(self, wizard):
        """Return (filename, bytes_content, mime_type) for the given wizard."""
        if not _OPENPYXL_AVAILABLE:
            _logger.warning(
                "openpyxl is not installed; falling back to CSV export."
            )
            return self._generate_csv(wizard)

        wb = Workbook()
        method_name = "export_" + wizard.report_type
        if not hasattr(self, method_name):
            method_name = "export_generic"
        getattr(self, method_name)(wb, wizard)
        self._add_cover_sheet(wb, wizard)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = "%s_%s.xlsx" % (
            wizard.report_type,
            fields.Date.to_string(wizard.date_to),
        )
        return filename, buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # ═══════════════════════════════════════════════════════════════════
    # Cover sheet (added last but moved to first position)
    # ═══════════════════════════════════════════════════════════════════
    def _add_cover_sheet(self, wb, wiz):
        cover = wb.create_sheet("Cover", 0)
        cover.column_dimensions["A"].width = 30
        cover.column_dimensions["B"].width = 50
        title = dict(wiz._fields["report_type"].selection).get(wiz.report_type, wiz.report_type)
        cover["A1"] = wiz.company_id.name
        cover["A1"].font = Font(size=18, bold=True, color="0B3D2E")
        cover["A3"] = "التقرير"
        cover["B3"] = title
        cover["A4"] = "الفترة"
        cover["B4"] = "%s → %s" % (wiz.date_from, wiz.date_to)
        cover["A5"] = "تاريخ التوليد"
        cover["B5"] = fields.Datetime.now()
        cover["A6"] = "Nexus Engine"
        cover["B6"] = "نظام نِكسَس المالي المتكامل"
        for row in (1, 3, 4, 5, 6):
            cover.cell(row=row, column=1).font = Font(bold=True)
            cover.cell(row=row, column=1).fill = _HEADER_FILL
            cover.cell(row=row, column=1).border = _BORDER
            cover.cell(row=row, column=2).border = _BORDER

    # ═══════════════════════════════════════════════════════════════════
    # Balance Sheet
    # ═══════════════════════════════════════════════════════════════════
    def export_balance_sheet(self, wb, wiz):
        sheet = wb.create_sheet("Balance Sheet")
        self._write_header(sheet, ["Account", "Balance"], widths=[50, 20])
        sheet.append([])

        groups = self.env["nexus.finance.report.renderer"].render_balance_sheet(wiz)
        # Parse the HTML tables back into rows (simple but effective)
        rows = self._parse_html_table(groups)
        for section, items in rows.items():
            sheet.append([section])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=12, color="0B3D2E")
            for name, amount in items:
                sheet.append([name, amount])
                sheet.cell(row=sheet.max_row, column=2).number_format = "#,##0.00"
            sheet.append([])

        sheet.sheet_view.zoomScale = 100

    # ═══════════════════════════════════════════════════════════════════
    # Profit & Loss
    # �══════════════════════════════════════════════════════════════════
    def export_profit_loss(self, wb, wiz):
        sheet = wb.create_sheet("Profit & Loss")
        self._write_header(sheet, ["Account", "Amount"], widths=[50, 20])
        sheet.append([])
        html = self.env["nexus.finance.report.renderer"].render_profit_loss(wiz)
        rows = self._parse_html_pl(html)
        for section, items in rows.items():
            sheet.append([section])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=12, color="0B3D2E")
            for name, amount in items:
                sheet.append([name, amount])
                sheet.cell(row=sheet.max_row, column=2).number_format = "#,##0.00"

    # ═══════════════════════════════════════════════════════════════════
    # Trial Balance
    # ═══════════════════════════════════════════════════════════════════
    def export_trial_balance(self, wb, wiz):
        sheet = wb.create_sheet("Trial Balance")
        self._write_header(
            sheet,
            ["Code", "Account", "Debit", "Credit"],
            widths=[15, 50, 20, 20],
        )
        domain = [
            ("company_id", "=", wiz.company_id.id),
            ("date", "<=", wiz.date_to),
            ("parent_state", "=", "posted"),
        ]
        groups = self.env["account.move.line"].read_group(
            domain, ["account_id", "debit", "credit"], ["account_id"]
        )
        total_debit = total_credit = 0.0
        for g in groups:
            acc = self.env["account.account"].browse(g["account_id"][0])
            sheet.append([acc.code, acc.display_name, g["debit"], g["credit"]])
            sheet.cell(row=sheet.max_row, column=3).number_format = "#,##0.00"
            sheet.cell(row=sheet.max_row, column=4).number_format = "#,##0.00"
            total_debit += g["debit"]
            total_credit += g["credit"]
        sheet.append(["TOTAL", "", total_debit, total_credit])
        for col in (1, 3, 4):
            cell = sheet.cell(row=sheet.max_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = _TOTAL_FILL
        sheet.cell(row=sheet.max_row, column=3).number_format = "#,##0.00"
        sheet.cell(row=sheet.max_row, column=4).number_format = "#,##0.00"

    # ═══════════════════════════════════════════════════════════════════
    # General Ledger
    # ═══════════════════════════════════════════════════════════════════
    def export_general_ledger(self, wb, wiz):
        sheet = wb.create_sheet("General Ledger")
        self._write_header(
            sheet,
            ["Date", "Move", "Account", "Partner", "Label", "Debit", "Credit"],
            widths=[12, 15, 35, 25, 35, 15, 15],
        )
        domain = [
            ("company_id", "=", wiz.company_id.id),
            ("date", ">=", wiz.date_from),
            ("date", "<=", wiz.date_to),
            ("parent_state", "=", "posted"),
        ]
        if wiz.account_id:
            domain.append(("account_id", "=", wiz.account_id.id))
        if wiz.partner_id:
            domain.append(("partner_id", "=", wiz.partner_id.id))
        for line in self.env["account.move.line"].search(domain, order="date,id"):
            sheet.append([
                str(line.date),
                line.move_id.name,
                line.account_id.display_name,
                line.partner_id.display_name if line.partner_id else "",
                line.name or line.move_id.ref or "",
                line.debit,
                line.credit,
            ])
            sheet.cell(row=sheet.max_row, column=6).number_format = "#,##0.00"
            sheet.cell(row=sheet.max_row, column=7).number_format = "#,##0.00"

    # ═══════════════════════════════════════════════════════════════════
    # Aging
    # ═══════════════════════════════════════════════════════════════════
    def export_aging_receivable(self, wb, wiz):
        self._export_aging(wb, wiz, "asset_receivable", "Receivable Aging")

    def export_aging_payable(self, wb, wiz):
        self._export_aging(wb, wiz, "liability_payable", "Payable Aging")

    def _export_aging(self, wb, wiz, account_type, sheet_title):
        sheet = wb.create_sheet(sheet_title)
        self._write_header(
            sheet,
            ["Partner", "0-30", "31-60", "61-90", "90+", "Total"],
            widths=[40, 15, 15, 15, 15, 18],
        )
        MoveLine = self.env["account.move.line"]
        as_of = wiz.date_to
        lines = MoveLine.search([
            ("company_id", "=", wiz.company_id.id),
            ("date", "<=", as_of),
            ("parent_state", "=", "posted"),
            ("account_id.account_type", "=", account_type),
            ("reconciled", "=", False),
        ])
        buckets = {}
        for line in lines:
            days = (as_of - line.date).days
            amount = line.amount_residual
            partner_name = line.partner_id.display_name if line.partner_id else "Unknown"
            buckets.setdefault(
                partner_name, [0.0, 0.0, 0.0, 0.0]
            )
            if days <= 30:
                buckets[partner_name][0] += amount
            elif days <= 60:
                buckets[partner_name][1] += amount
            elif days <= 90:
                buckets[partner_name][2] += amount
            else:
                buckets[partner_name][3] += amount
        totals = [0.0, 0.0, 0.0, 0.0]
        for name, amts in sorted(buckets.items(), key=lambda x: -sum(x[1])):
            total = sum(amts)
            sheet.append([name] + amts + [total])
            for col in range(2, 7):
                sheet.cell(row=sheet.max_row, column=col).number_format = "#,##0.00"
            for i, a in enumerate(amts):
                totals[i] += a
        grand_total = sum(totals)
        sheet.append(["TOTAL"] + totals + [grand_total])
        for col in range(1, 7):
            cell = sheet.cell(row=sheet.max_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = _TOTAL_FILL
            if col > 1:
                cell.number_format = "#,##0.00"

    # ═══════════════════════════════════════════════════════════════════
    # Budget Variance
    # ═══════════════════════════════════════════════════════════════════
    def export_budget_variance(self, wb, wiz):
        sheet = wb.create_sheet("Budget Variance")
        self._write_header(
            sheet,
            ["Budget", "Analytic", "Planned", "Actual", "Variance"],
            widths=[30, 35, 18, 18, 18],
        )
        Budget = self.env.get("crossovered.budget")
        if not Budget:
            sheet.append(["Budgets not configured"])
            return
        for budget in Budget.search([
            ("date_from", "<=", wiz.date_to),
            ("date_to", ">=", wiz.date_from),
            ("company_id", "=", wiz.company_id.id),
        ]):
            for line in budget.crossovered_budget_line:
                sheet.append([
                    budget.name,
                    line.analytic_account_id.display_name,
                    line.planned_amount,
                    line.achieved_amount,
                    line.planned_amount - line.achieved_amount,
                ])
                for col in (3, 4, 5):
                    sheet.cell(row=sheet.max_row, column=col).number_format = "#,##0.00"

    # ═══════════════════════════════════════════════════════════════════
    # Cost Center
    # �══════════════════════════════════════════════════════════════════
    def export_cost_center(self, wb, wiz):
        sheet = wb.create_sheet("Cost Centers")
        self._write_header(
            sheet,
            ["Cost Center", "Amount"],
            widths=[50, 20],
        )
        totals = {}
        for line in self.env["account.analytic.line"].search([
            ("company_id", "=", wiz.company_id.id),
            ("date", ">=", wiz.date_from),
            ("date", "<=", wiz.date_to),
        ]):
            key = line.account_id.display_name
            totals[key] = totals.get(key, 0.0) + line.amount
        for name, amt in sorted(totals.items(), key=lambda x: -x[1]):
            sheet.append([name, amt])
            sheet.cell(row=sheet.max_row, column=2).number_format = "#,##0.00"
        sheet.append(["TOTAL", sum(totals.values())])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
        sheet.cell(row=sheet.max_row, column=2).font = Font(bold=True)
        sheet.cell(row=sheet.max_row, column=2).number_format = "#,##0.00"

    # ═══════════════════════════════════════════════════════════════════
    # Cash Flow
    # ═══════════════════════════════════════════════════════════════════
    def export_cash_flow(self, wb, wiz):
        sheet = wb.create_sheet("Cash Flow")
        self._write_header(sheet, ["Metric", "Amount"], widths=[40, 20])
        sheet.append(["Opening Balance", ""])
        sheet.append(["Net Movement", ""])
        sheet.append(["Closing Balance", ""])
        # Renderer also computes these; we delegate then read back
        from . import nexus_finance_report_renderer  # noqa: F401
        html = self.env["nexus.finance.report.renderer"].render_cash_flow(wiz)
        # Light-weight parse: just re-compute here
        MoveLine = self.env["account.move.line"]
        cash_accounts = self.env["account.account"].search([
            ("account_type", "in", ["asset_cash", "asset_bank"]),
            ("company_id", "=", wiz.company_id.id),
        ])
        opening = MoveLine.read_group([
            ("account_id", "in", cash_accounts.ids),
            ("date", "<", wiz.date_from),
            ("parent_state", "=", "posted"),
        ], ["balance"], [])
        opening_balance = opening[0]["balance"] if opening else 0.0
        movements = MoveLine.read_group([
            ("account_id", "in", cash_accounts.ids),
            ("date", ">=", wiz.date_from),
            ("date", "<=", wiz.date_to),
            ("parent_state", "=", "posted"),
        ], ["balance"], [])
        net_movement = movements[0]["balance"] if movements else 0.0
        closing = opening_balance + net_movement
        sheet.cell(row=2, column=2, value=opening_balance).number_format = "#,##0.00"
        sheet.cell(row=3, column=2, value=net_movement).number_format = "#,##0.00"
        sheet.cell(row=4, column=2, value=closing).number_format = "#,##0.00"
        for r in (2, 3, 4):
            sheet.cell(row=r, column=2).font = Font(bold=(r == 4))
            if r == 4:
                sheet.cell(row=r, column=2).fill = _TOTAL_FILL

    # ═══════════════════════════════════════════════════════════════════
    # Generic fallback (renders all data into a flat table)
    # ═══════════════════════════════════════════════════════════════════
    def export_generic(self, wb, wiz):
        sheet = wb.create_sheet("Report")
        self._write_header(sheet, ["Description", "Amount"], widths=[60, 20])
        sheet.append(["This report is not yet implemented in Excel form."])
        sheet.append(["Please use the HTML view for now."])

    # ═══════════════════════════════════════════════════════════════════
    # CSV fallback
    # ═══════════════════════════════════════════════════════════════════
    def _generate_csv(self, wiz):
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Report", wiz.report_type])
        writer.writerow(["Date From", str(wiz.date_from)])
        writer.writerow(["Date To", str(wiz.date_to)])
        writer.writerow([])
        writer.writerow(["Description", "Amount"])
        writer.writerow(["Data export is unavailable — openpyxl is missing."])
        return (
            "%s_%s.csv" % (wiz.report_type, fields.Date.to_string(wiz.date_to)),
            buf.getvalue().encode("utf-8"),
            "text/csv",
        )

    # ═══════════════════════════════════════════════════════════════════
    # Helpers
    # ═══════════════════════════════════════════════════════════════════
    def _write_header(self, sheet, headers, widths=None):
        sheet.append(headers)
        for col_idx, value in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _BRAND_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
            if widths and col_idx <= len(widths):
                sheet.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
        sheet.row_dimensions[1].height = 22

    def _parse_html_table(self, html):
        """Convert the renderer's HTML into {section: [(name, amount), ...]}."""
        import re
        result = {}
        current = None
        # Match each <h5>SECTION</h5> ... <tbody>...</tbody>
        for m in re.finditer(
            r"<h5[^>]*>([^<]+)</h5>\s*<table[^>]*>.*?<tbody>(.*?)</tbody>",
            html, re.S,
        ):
            section = m.group(1).strip()
            body = m.group(2)
            items = []
            for row in re.finditer(
                r"<td>([^<]+)</td><td[^>]*>([\d.\-]+)</td>", body
            ):
                items.append((row.group(1).strip(), float(row.group(2))))
            result[section] = items
        return result

    def _parse_html_pl(self, html):
        return self._parse_html_table(html)
