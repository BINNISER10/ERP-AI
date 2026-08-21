"""Nexus Setup Journey: a gamified, stage-based onboarding engine.

The journey walks the user through five critical setup stages and silently
pushes each completed stage to the Nexus Core via an outbound API call.
"""
import json
import logging
import requests

from odoo import api, fields, models, _
from odoo.exceptions import UserError, RedirectWarning

_logger = logging.getLogger(__name__)

STAGES = [
    ("identity_legal", "Identity & Legal"),
    ("structure_hr", "Structure & HR"),
    ("financial_core", "Financial Core"),
    ("operations", "Operations"),
    ("go_live", "Go-Live"),
]

STEP_CODES = {
    "identity_legal": {
        "label": _("Identity & Legal"),
        "hint": _("Company name, logo, tax ID, currency and industry."),
        "sequence": 10,
    },
    "structure_hr": {
        "label": _("Structure & HR"),
        "hint": _("Branches/warehouses, departments and employees."),
        "sequence": 20,
    },
    "financial_core": {
        "label": _("Financial Core"),
        "hint": _("Bank accounts, cash registers, payment methods and taxes."),
        "sequence": 30,
    },
    "operations": {
        "label": _("Operations"),
        "hint": _("Product categories, items and POS terminal."),
        "sequence": 40,
    },
    "go_live": {
        "label": _("Go-Live"),
        "hint": _("Final validation and readiness check."),
        "sequence": 50,
    },
}


class NexusSetupStep(models.Model):
    """A single step inside a Nexus setup journey."""

    _name = "nexus.setup.step"
    _description = "Nexus Setup Step"
    _order = "sequence, id"

    journey_id = fields.Many2one(
        "nexus.setup.journey",
        string="Journey",
        required=True,
        ondelete="cascade",
    )
    name = fields.Char(required=True, translate=True)
    code = fields.Char(required=True)
    stage = fields.Selection(STAGES, required=True)
    sequence = fields.Integer(default=10)
    description = fields.Text(translate=True)
    required = fields.Boolean(default=True)
    state = fields.Selection(
        [
            ("pending", "Pending"),
            ("in_progress", "In Progress"),
            ("done", "Done"),
            ("skipped", "Skipped"),
        ],
        default="pending",
        required=True,
    )
    auto_sync = fields.Boolean(
        default=True,
        help="If set, completing this step will push the configuration to Nexus Core.",
    )
    sync_payload = fields.Text(
        string="Last Sync Payload",
        help="JSON snapshot sent to Nexus Core.",
    )
    last_sync_status = fields.Selection(
        [
            ("pending", "Pending"),
            ("success", "Success"),
            ("warning", "Warning"),
            ("error", "Error"),
        ],
        default="pending",
    )
    last_sync_message = fields.Text()
    last_sync_date = fields.Datetime()

    def action_open_form(self):
        """Return the journey form view focused on this step."""
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "res_model": self.journey_id._name,
            "res_id": self.journey_id.id,
            "view_mode": "form",
            "target": "current",
        }


def _normalize_arabic_digits(text):
    """Normalize Arabic-Indic numerals (Ù Ù¡Ù¢Ù£Ù¤Ù¥Ù¦Ù§Ù¨Ù©) to standard ASCII digits (0-9)."""
    if not text:
        return ""
    trans = str.maketrans("Ù Ù¡Ù¢Ù£Ù¤Ù¥Ù¦Ù§Ù¨Ù©", "0123456789")
    return text.translate(trans)


class NexusScannedDocument(models.Model):
    """Stores individual scanned documents and their extracted metadata in batch processing."""
    _name = "nexus.scanned.document"
    _description = "Nexus Scanned Business Document"
    _order = "id desc"

    journey_id = fields.Many2one("nexus.setup.journey", string="Setup Journey", ondelete="cascade")
    name = fields.Char(string="Ø§Ø³Ù… Ø§Ù„Ù…Ù„Ù (File Name)", required=True)
    attachment_id = fields.Many2one("ir.attachment", string="Ø§Ù„Ù…Ù„Ù Ø§Ù„Ù…Ø±ÙÙ‚")
    file_type = fields.Char(string="Ù†ÙˆØ¹ Ø§Ù„Ù…Ù„Ù")
    document_type = fields.Selection(
        [
            ("cr", "ðŸ“‘ Ø§Ù„Ø³Ø¬Ù„ Ø§Ù„ØªØ¬Ø§Ø±ÙŠ"),
            ("vat", "ðŸ§¾ Ø´Ù‡Ø§Ø¯Ø© Ø¶Ø±ÙŠØ¨Ø© Ø§Ù„Ù‚ÙŠÙ…Ø© Ø§Ù„Ù…Ø¶Ø§ÙØ©"),
            ("national_address", "ðŸ“ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„ÙˆØ·Ù†ÙŠ"),
            ("gosi", "ðŸ›¡ï¸ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„ØªØ£Ù…ÙŠÙ†Ø§Øª Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ©"),
            ("balady", "ðŸ¢ Ø±Ø®ØµØ© Ø§Ù„Ù†Ø´Ø§Ø· Ø§Ù„ØªØ¬Ø§Ø±ÙŠ (Ø¨Ù„Ø¯ÙŠ)"),
            ("nitaqat", "ðŸ‘¥ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„Ø³Ø¹ÙˆØ¯Ø© ÙˆÙ†Ø·Ø§Ù‚Ø§Øª"),
            ("chamber", "ðŸ›ï¸ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„ØºØ±ÙØ© Ø§Ù„ØªØ¬Ø§Ø±ÙŠØ©"),
            ("invoice", "ðŸ’° ÙØ§ØªÙˆØ±Ø© / Ø³Ù†Ø¯ Ù…Ø§Ù„ÙŠ"),
            ("contract", "ðŸ“ Ø¹Ù‚Ø¯ ØªØ¬Ø§Ø±ÙŠ"),
            ("other", "ðŸ“„ ÙˆØ«ÙŠÙ‚Ø© Ø£Ø¹Ù…Ø§Ù„ Ø¹Ø§Ù…Ø©"),
        ],
        string="Ù†ÙˆØ¹ Ø§Ù„ÙˆØ«ÙŠÙ‚Ø© Ø§Ù„Ù…ØµØ·Ø§Ø¯Ø©",
        default="other",
    )
    extracted_summary = fields.Char(string="Ø§Ù„Ù…Ù„Ø®Øµ Ø§Ù„Ù…Ø³ØªØ®Ø±Ø¬")
    extracted_data_json = fields.Text(string="Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù…Ø³ØªØ®Ø±Ø¬Ø© (JSON)")
    state = fields.Selection(
        [
            ("uploaded", "Ø¬Ø§Ù‡Ø² Ù„Ù„ÙØ­Øµ"),
            ("extracted", "ØªÙ… Ø§Ù„ØµÙŠØ¯ ÙˆØ§Ù„Ø§Ø³ØªØ®Ø±Ø§Ø¬ âœ…"),
            ("failed", "ØªØ¹Ø°Ø± Ø§Ù„Ø§Ø³ØªØ®Ø±Ø§Ø¬ âš ï¸"),
        ],
        string="Ø§Ù„Ø­Ø§Ù„Ø©",
        default="uploaded",
    )


class NexusSetupJourney(models.Model):
    """Gamified onboarding journey for a company."""

    _name = "nexus.setup.journey"
    _description = "Nexus Setup Journey"
    _rec_name = "display_name"
    _order = "id desc"
    _inherit = ["mail.thread"]

    name = fields.Char(compute="_compute_name", store=True)
    display_name = fields.Char(compute="_compute_display_name", store=True)
    company_id = fields.Many2one(
        "res.company",
        string="Company",
        required=True,
        default=lambda self: self.env.company,
        index=True,
    )
    active = fields.Boolean(default=True)

    stage = fields.Selection(
        STAGES + [("done", "Done")],
        default="identity_legal",
        required=True,
        tracking=True,
    )
    state = fields.Selection(
        [
            ("draft", "Draft"),
            ("in_progress", "In Progress"),
            ("done", "Done"),
            ("cancelled", "Cancelled"),
        ],
        default="draft",
        required=True,
    )
    progress = fields.Integer(
        string="Progress (%)",
        compute="_compute_progress",
        store=True,
        default=0,
    )

    # Identity & Legal
    company_name = fields.Char(
        related="company_id.name",
        string="Company Name",
        readonly=False,
    )
    company_logo = fields.Binary(
        related="company_id.logo",
        string="Company Logo",
        readonly=False,
    )
    tax_id = fields.Char(
        related="company_id.vat",
        string="Tax ID",
        readonly=False,
    )
    currency_id = fields.Many2one(
        "res.currency",
        related="company_id.currency_id",
        string="Currency",
        readonly=False,
    )
    industry_domain = fields.Selection(
        [
            ("retail", "Retail"),
            ("restaurant", "Restaurant / F&B"),
            ("manufacturing", "Manufacturing"),
            ("construction", "Construction"),
            ("services", "Services"),
            ("healthcare", "Healthcare"),
            ("education", "Education"),
            ("logistics", "Logistics"),
            ("fuel_station", "Fuel Station"),
            ("real_estate", "Real Estate"),
            ("other", "Other"),
        ],
        string="Industry / Domain",
        default="services",
    )

    # Document Hunter Fields (Multi-File Dropzone up to 100+ documents)
    document_attachment_ids = fields.Many2many(
        "ir.attachment",
        "nexus_journey_attachment_rel",
        "journey_id",
        "attachment_id",
        string="Ø§Ù„Ù…Ø³ØªÙ†Ø¯Ø§Øª Ø§Ù„Ù…Ø±ÙÙˆØ¹Ø© (Ø­ØªÙ‰ 100+ ÙˆØ«ÙŠÙ‚Ø©)",
    )
    scanned_document_ids = fields.One2many(
        "nexus.scanned.document",
        "journey_id",
        string="Ø³Ø¬Ù„ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ù…ÙØ­ÙˆØµØ© ÙˆØ§Ù„Ù…ØµØ·Ø§Ø¯Ø©",
    )
    total_documents_count = fields.Integer(
        string="Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚",
        compute="_compute_document_counts",
    )
    scanned_documents_count = fields.Integer(
        string="Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ù…Ø¹Ø§Ù„Ø¬Ø©",
        compute="_compute_document_counts",
    )

    upload_document_file = fields.Binary(string="Ø§Ø³Ø­Ø¨ Ø£Ùˆ Ø§Ø±ÙØ¹ Ø£ÙŠ ÙˆØ«ÙŠÙ‚Ø© Ø±Ø³Ù…ÙŠØ© Ù‡Ù†Ø§ (Single AI Dropzone)", attachment=True)
    upload_document_filename = fields.Char(string="Ø§Ø³Ù… Ù…Ù„Ù Ø§Ù„ÙˆØ«ÙŠÙ‚Ø©")
    
    # Saudi Official Registry & Licenses
    cr_number = fields.Char(string="Ø±Ù‚Ù… Ø§Ù„Ø³Ø¬Ù„ Ø§Ù„ØªØ¬Ø§Ø±ÙŠ", placeholder="1010XXXXXX")
    gosi_number = fields.Char(string="Ø±Ù‚Ù… Ø§Ø´ØªØ±Ø§Ùƒ Ø§Ù„ØªØ£Ù…ÙŠÙ†Ø§Øª", placeholder="700XXXXXXX")
    balady_license_no = fields.Char(string="Ø±Ù‚Ù… Ø±Ø®ØµØ© Ø¨Ù„Ø¯ÙŠ", placeholder="1445XXXXXXXX")
    saudization_rate = fields.Float(string="Ù†Ø³Ø¨Ø© Ø§Ù„ØªÙˆØ·ÙŠÙ† / Ø§Ù„Ø³Ø¹ÙˆØ¯Ø© (%)")

    # National Address Fields
    city = fields.Char(string="Ø§Ù„Ù…Ø¯ÙŠÙ†Ø© (City)", default="Ø§Ù„Ø±ÙŠØ§Ø¶")
    district = fields.Char(string="Ø§Ù„Ø­ÙŠ (District)")
    street = fields.Char(string="Ø§Ù„Ø´Ø§Ø±Ø¹ (Street)")
    building_no = fields.Char(string="Ø±Ù‚Ù… Ø§Ù„Ù…Ø¨Ù†Ù‰ (Building No)")
    additional_no = fields.Char(string="Ø§Ù„Ø±Ù‚Ù… Ø§Ù„Ø¥Ø¶Ø§ÙÙŠ (Additional No)")
    postal_code = fields.Char(string="Ø§Ù„Ø±Ù…Ø² Ø§Ù„Ø¨Ø±ÙŠØ¯ÙŠ (Postal Code)")
    national_short_address = fields.Char(string="Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„Ù…Ø®ØªØµØ±", placeholder="RRRD2934")

    employee_count = fields.Integer(string="Ø¹Ø¯Ø¯ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†", default=5)
    document_scan_summary_html = fields.Html(string="Scan Summary", readonly=True)
    recognized_documents_html = fields.Html(string="Ø³Ø¬Ù„ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ù…ØµØ·Ø§Ø¯Ø©", readonly=True)

    # HR Smart Importer & Employee Document Hunter Fields
    hr_roster_file = fields.Binary(string="ÙƒØ´Ù Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† (Excel / CSV / PDF)", attachment=True)
    hr_roster_filename = fields.Char(string="Ø§Ø³Ù… Ù…Ù„Ù ÙƒØ´Ù Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†")
    employee_document_attachment_ids = fields.Many2many(
        "ir.attachment",
        "nexus_journey_emp_att_rel",
        "journey_id",
        "attachment_id",
        string="ØµÙˆØ± ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† (Ø¥Ù‚Ø§Ù…Ø§Øª ÙˆØ¬ÙˆØ§Ø²Ø§Øª Ù…ØªØ¹Ø¯Ø¯Ø©)",
    )
    hr_scan_summary_html = fields.Html(string="Ù…Ù„Ø®Øµ Ø§Ø³ØªÙŠØ±Ø§Ø¯ ÙˆÙØ­Øµ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†", readonly=True)

    # Legacy fields for backward compatibility
    cr_file = fields.Binary(string="Ø§Ù„Ø³Ø¬Ù„ Ø§Ù„ØªØ¬Ø§Ø±ÙŠ (CR)", attachment=True)
    cr_filename = fields.Char(string="CR Filename")
    vat_file = fields.Binary(string="Ø§Ù„Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„Ø¶Ø±ÙŠØ¨ÙŠØ© (VAT)", attachment=True)
    vat_filename = fields.Char(string="VAT Filename")
    gosi_file = fields.Binary(string="Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„ØªØ£Ù…ÙŠÙ†Ø§Øª (GOSI)", attachment=True)
    gosi_filename = fields.Char(string="GOSI Filename")

    # Structure & HR
    warehouse_ids = fields.Many2many(
        "stock.warehouse",
        string="Branches / Warehouses",
        help="Physical locations that act as branches or warehouses.",
    )
    department_ids = fields.Many2many(
        "hr.department",
        string="Departments",
    )
    employee_ids = fields.Many2many(
        "hr.employee",
        string="Employees",
    )

    # Financial Core
    bank_journal_ids = fields.Many2many(
        "account.journal",
        "journey_bank_journal_rel",
        "journey_id",
        "journal_id",
        string="Bank Accounts",
        domain="[('type', '=', 'bank'), ('company_id', '=', company_id)]",
    )
    cash_journal_ids = fields.Many2many(
        "account.journal",
        "journey_cash_journal_rel",
        "journey_id",
        "journal_id",
        string="Cash Registers",
        domain="[('type', '=', 'cash'), ('company_id', '=', company_id)]",
    )
    payment_method_ids = fields.Many2many(
        "pos.payment.method",
        string="Payment Methods",
    )
    tax_ids = fields.Many2many(
        "account.tax",
        string="Taxes",
        domain="[('company_id', '=', company_id)]",
    )

    # Operations
    product_category_ids = fields.Many2many(
        "product.category",
        string="Product Categories",
    )
    product_ids = fields.Many2many(
        "product.template",
        string="Items / Products",
        domain="[('company_id', '=', company_id)]",
    )
    pos_config_ids = fields.Many2many(
        "pos.config",
        string="POS Terminals",
    )

    step_ids = fields.One2many(
        "nexus.setup.step",
        "journey_id",
        string="Setup Steps",
    )
    notes = fields.Text(string="Setup Notes")

    _sql_constraints = [
        (
            "company_uniq",
            "unique(company_id)",
            "Only one setup journey is allowed per company.",
        ),
    ]

    @api.depends("company_id")
    def _compute_name(self):
        for journey in self:
            journey.name = journey.company_id.name or _("New Journey")

    @api.depends("name", "progress", "stage")
    def _compute_display_name(self):
        for journey in self:
            label = dict(journey._fields["stage"].selection).get(journey.stage, journey.stage)
            journey.display_name = f"{journey.name} â€” {label} ({journey.progress}%)"

    @api.depends("step_ids.state", "step_ids.required")
    def _compute_progress(self):
        for journey in self:
            steps = journey.step_ids
            required = steps.filtered(lambda s: s.required)
            if not required:
                journey.progress = 0
                continue
            done = required.filtered(lambda s: s.state in ("done", "skipped"))
            journey.progress = int(round(len(done) / len(required) * 100))

    @api.model_create_multi
    def create(self, vals_list):
        journeys = super().create(vals_list)
        for journey in journeys:
            journey._create_default_steps()
        return journeys

    def _create_default_steps(self):
        self.ensure_one()
        Step = self.env["nexus.setup.step"]
        existing = set(self.step_ids.mapped("code"))
        to_create = []
        for code, meta in STEP_CODES.items():
            if code in existing:
                continue
            to_create.append({
                "journey_id": self.id,
                "name": meta["label"],
                "code": code,
                "stage": code,
                "sequence": meta["sequence"],
                "description": meta["hint"],
            })
        if to_create:
            Step.create(to_create)

    @api.model
    def get_or_create(self, company=None):
        """Return the journey for the given company, creating it if needed."""
        company = company or self.env.company
        journey = self.search([("company_id", "=", company.id)], limit=1)
        if not journey:
            journey = self.create({"company_id": company.id})
        return journey

    def action_start(self):
        """Move from draft to in_progress."""
        self.write({"state": "in_progress"})

    def _get_step(self, code):
        self.ensure_one()
        return self.step_ids.filtered(lambda s: s.code == code)[:1]

    def action_mark_step_done(self, step_code=None):
        """Mark the requested step done and (optionally) sync it to Nexus Core."""
        self.ensure_one()
        step_code = step_code or self.env.context.get("step_code") or self.stage
        step = self._get_step(step_code)
        if not step:
            raise UserError(_("Step %s not found in this journey.", step_code))

        if self.state == "draft":
            self.state = "in_progress"

        step.state = "done"
        step.last_sync_date = fields.Datetime.now()

        if step.auto_sync:
            self.action_sync_step_to_core(step)

        self._advance_stage()
        return self._reopen_form()

    def _advance_stage(self):
        """Set the current stage to the next incomplete required step."""
        self.ensure_one()
        for code, meta in STEP_CODES.items():
            step = self._get_step(code)
            if step and step.state != "done" and step.required:
                self.stage = code
                return
        self.stage = "done"
        if self.state != "done":
            self.state = "in_progress"

    def _get_nexus_core_config(self):
        """Return the active Copilot/Hybrid config or an empty recordset."""
        self.ensure_one()
        config = self.env["copilot.config"].sudo().get_active_config(self.company_id)
        if not config or not config.nexus_core_url:
            return self.env["hybrid.config"].sudo().get_active_config(self.company_id)
        return config

    def _build_stage_payload(self, step):
        """Build a JSON payload that describes the current stage configuration."""
        self.ensure_one()
        payload = {
            "journey_id": self.id,
            "company_id": self.company_id.id,
            "company_name": self.company_id.name,
            "stage": step.stage,
            "step_name": step.name,
            "timestamp": fields.Datetime.now().isoformat(),
            "config": {},
        }
        if step.stage == "identity_legal":
            payload["config"] = {
                "company_name": self.company_name,
                "tax_id": self.tax_id,
                "currency": self.currency_id.name if self.currency_id else None,
                "industry_domain": self.industry_domain,
            }
        elif step.stage == "structure_hr":
            payload["config"] = {
                "warehouses": [
                    {"id": w.id, "name": w.name, "code": w.code}
                    for w in self.warehouse_ids
                ],
                "departments": [
                    {"id": d.id, "name": d.name}
                    for d in self.department_ids
                ],
                "employees": [
                    {"id": e.id, "name": e.name}
                    for e in self.employee_ids
                ],
            }
        elif step.stage == "financial_core":
            payload["config"] = {
                "bank_journals": [j.name for j in self.bank_journal_ids],
                "cash_journals": [j.name for j in self.cash_journal_ids],
                "payment_methods": [m.name for m in self.payment_method_ids],
                "taxes": [t.name for t in self.tax_ids],
            }
        elif step.stage == "operations":
            payload["config"] = {
                "product_categories": [c.name for c in self.product_category_ids],
                "products": [p.name for p in self.product_ids],
                "pos_configs": [c.name for c in self.pos_config_ids],
            }
        elif step.stage == "go_live":
            payload["config"] = {
                "progress": self.progress,
                "state": self.state,
                "ready": self.progress >= 100,
            }
        return payload

    def action_sync_step_to_core(self, step=None):
        """Silently push the given step configuration to the Nexus Core API."""
        self.ensure_one()
        if step is None:
            step = self._get_step(self.stage)
        if not step:
            return False

        config = self._get_nexus_core_config()
        if not config or not getattr(config, "nexus_core_url", False):
            step.write({
                "last_sync_status": "warning",
                "last_sync_message": _("Nexus Core URL is not configured."),
            })
            return False

        url = config.nexus_core_url.rstrip("/") + "/api/v1/setup/sync"
        payload = self._build_stage_payload(step)
        step.sync_payload = json.dumps(payload, indent=2, default=str)

        headers = {"Content-Type": "application/json"}
        api_key = getattr(config, "nexus_core_api_key", False) or getattr(config, "n8n_webhook_key", False)
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"

        try:
            response = requests.post(url, json=payload, headers=headers, timeout=15)
            response.raise_for_status()
            step.write({
                "last_sync_status": "success",
                "last_sync_message": _(
                    "Sync successful (HTTP %s).", response.status_code
                ),
                "last_sync_date": fields.Datetime.now(),
            })
        except requests.exceptions.RequestException as exc:
            _logger.exception("Nexus Core sync failed for step %s", step.code)
            step.write({
                "last_sync_status": "error",
                "last_sync_message": str(exc),
                "last_sync_date": fields.Datetime.now(),
            })
        except Exception as exc:
            _logger.exception("Unexpected sync error for step %s", step.code)
            step.write({
                "last_sync_status": "error",
                "last_sync_message": str(exc),
                "last_sync_date": fields.Datetime.now(),
            })

        return True

    def action_sync_all(self):
        """Sync every pending and done step to Nexus Core."""
        for journey in self:
            for step in journey.step_ids:
                journey.action_sync_step_to_core(step)

    def action_validate_go_live(self):
        """Run final validation and mark the journey done if ready."""
        self.ensure_one()
        if self.progress < 100:
            raise UserError(_(
                "You must complete all setup stages before going live."
            ))
        go_live = self._get_step("go_live")
        if go_live:
            go_live.state = "done"
        self.write({"stage": "done", "state": "done"})
        self.action_sync_step_to_core(go_live)
        return self._reopen_form()

    def action_open_setup(self):
        """Return an action that opens this journey's form view."""
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "current",
            "context": self.env.context,
        }

    def _reopen_form(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "current",
            "context": self.env.context,
        }

    # Flexible enforcement / warm reminders
    def _get_warm_message(self, scenario):
        """Return a warm, AI-friendly reminder or None if everything is ready."""
        self.ensure_one()
        if scenario == "pos_session":
            if not self.payment_method_ids:
                return _(
                    "You're almost there! You've set up your products beautifully, "
                    "but we just need to add a Payment Method so you can collect money. "
                    "Click here to add one."
                )
            if not self.pos_config_ids:
                return _(
                    "Almost ready! Please configure a POS terminal first so we can "
                    "open a session for you."
                )
        return None

    def action_check_pos_readiness(self):
        """Return an action or a friendly RedirectWarning for POS.

        This is a warm reminder, not a hard block.
        """
        self.ensure_one()
        msg = self._get_warm_message("pos_session")
        if msg:
            raise RedirectWarning(
                msg,
                {
                    "type": "ir.actions.act_window",
                    "res_model": self._name,
                    "res_id": self.id,
                    "views": [[False, "form"]],
                    "target": "current",
                },
                _("Open Setup Journey"),
            )
        return {"type": "ir.actions.act_window.close_all"}

    def action_jumpstart_onboarding(self):
        """Launch the Smart Onboarding Wizard for this journey."""
        self.ensure_one()
        return self.env["nexus.onboarding.wizard"].jump_start(self.company_id)

    def action_open_document_hunter_wizard(self):
        """Open the AI Document Hunter Wizard."""
        return {
            "name": "ðŸŽ¯ ØµÙŠØ§Ø¯ ÙˆØ§Ø³ØªØ®Ø±Ø§Ø¬ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚ Ø¨Ø§Ù„Ø°ÙƒØ§Ø¡ Ø§Ù„Ø§ØµØ·Ù†Ø§Ø¹ÙŠ",
            "type": "ir.actions.act_window",
            "res_model": "nexus.document.hunter.wizard",
            "view_mode": "form",
            "target": "new",
        }

    def action_open_ai_developer(self):
        """Open the AI Developer Staff Member."""
        return {
            "name": "ðŸ‘¨â€ðŸ’» Ù…Ø·ÙˆØ± Ø£ÙˆØ¯Ùˆ ÙˆÙ…Ø³ØªØ´Ø§Ø± Ø§Ù„Ø£Ø¹Ù…Ø§Ù„ Ø§Ù„Ø°ÙƒÙŠ",
            "type": "ir.actions.act_window",
            "res_model": "nexus.ai.developer.staff",
            "view_mode": "form",
            "target": "new",
        }

    @api.depends("document_attachment_ids", "scanned_document_ids")
    def _compute_document_counts(self):
        for j in self:
            j.total_documents_count = len(j.document_attachment_ids)
            j.scanned_documents_count = len(j.scanned_document_ids.filtered(lambda d: d.state == "extracted"))

    def action_clear_document_inputs(self):
        """Clear and reset all uploaded attachments, scanned records, and scan summaries."""
        self.ensure_one()
        self.document_attachment_ids = [(5, 0, 0)]
        self.upload_document_file = False
        self.upload_document_filename = False
        self.cr_file = False
        self.cr_filename = False
        self.vat_file = False
        self.vat_filename = False
        self.gosi_file = False
        self.gosi_filename = False
        if self.scanned_document_ids:
            self.scanned_document_ids.unlink()
        self.document_scan_summary_html = False

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("ðŸ§¹ ØªÙ… Ù…Ø³Ø­ Ø§Ù„Ù…Ø¯Ø®Ù„Ø§Øª"),
                "message": _("ØªÙ… ØªÙØ±ÙŠØº Ù…Ø±Ø¨Ø¹ Ø§Ù„Ø±ÙØ¹ ÙˆØ³Ø¬Ù„ Ø§Ù„ÙØ­Øµ Ø¨Ù†Ø¬Ø§Ø­. ÙŠÙ…ÙƒÙ†Ùƒ Ø§Ù„Ø¢Ù† Ø³Ø­Ø¨ ÙˆØ¥ÙÙ„Ø§Øª Ù…Ù„ÙØ§Øª Ø¬Ø¯ÙŠØ¯Ø©."),
                "type": "info",
                "sticky": False,
            },
        }

    def action_trigger_document_hunter(self):
        """Batch Multi-Document AI Hunter â€” Scans and auto-classifies 1 to 100+ documents at once!"""
        self.ensure_one()
        import base64
        import re

        files_to_process = []

        # 1. Collect from Multi-File attachments (supports 100+ files)
        for att in self.document_attachment_ids:
            if att.datas:
                files_to_process.append({
                    "name": att.name or "document.pdf",
                    "data": att.datas,
                    "attachment": att,
                })

        # 2. Collect from single dropzone if present
        if self.upload_document_file:
            files_to_process.append({
                "name": self.upload_document_filename or "document.pdf",
                "data": self.upload_document_file,
                "attachment": False,
            })

        # 3. Fallback to legacy fields
        for b_data, f_name in [(self.cr_file, self.cr_filename), (self.vat_file, self.vat_filename), (self.gosi_file, self.gosi_filename)]:
            if b_data:
                files_to_process.append({"name": f_name or "document.pdf", "data": b_data, "attachment": False})

        if not files_to_process:
            raise UserError(_("ÙŠØ±Ø¬Ù‰ Ø³Ø­Ø¨ Ø£Ùˆ Ø±ÙØ¹ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ø±Ø³Ù…ÙŠØ© ÙÙŠ Ù…Ø±Ø¨Ø¹ Ø§Ù„Ø±ÙØ¹ Ø§Ù„Ù…ÙˆØ­Ø¯ Ø£ÙˆÙ„Ø§Ù‹ Ù„Ø¨Ø¯Ø¡ Ø§Ù„ÙØ­Øµ ÙˆØ§Ù„ØµÙŠØ¯ Ø§Ù„Ø¢Ù„ÙŠ."))

        scanned_doc_obj = self.env["nexus.scanned.document"]
        processed_count = 0
        detected_types_summary = []

        for item in files_to_process:
            raw_bytes = None
            try:
                raw_bytes = base64.b64decode(item["data"])
            except Exception:
                continue

            raw_txt = raw_bytes.decode("utf-8", errors="ignore")
            txt = _normalize_arabic_digits(raw_txt)
            full_lower = txt.lower()

            # 1. CR Number (10 digits starting with 1 to 5)
            cr_m = re.search(r"(?:Ø³Ø¬Ù„\s*ØªØ¬Ø§Ø±ÙŠ|Ø±Ù‚Ù…\s*Ø§Ù„Ø³Ø¬Ù„|cr\s*no|commercial\s*registration)[\s:â€“-]*([1-5]\d{9})", txt, re.I)
            if not cr_m:
                cr_m = re.search(r"\b([1-5]\d{9})\b", txt)
            if cr_m:
                self.cr_number = cr_m.group(1)

            # 2. VAT Number (15 digits starting and ending with 3)
            vat_m = re.search(r"(?:Ø§Ù„Ø±Ù‚Ù…\s*Ø§Ù„Ø¶Ø±ÙŠØ¨ÙŠ|Ø¶Ø±ÙŠØ¨Ø©\s*Ø§Ù„Ù‚ÙŠÙ…Ø©\s*Ø§Ù„Ù…Ø¶Ø§ÙØ©|Ø±Ù‚Ù…\s*Ø§Ù„ØªØ³Ø¬ÙŠÙ„\s*Ø§Ù„Ø¶Ø±ÙŠØ¨ÙŠ|vat\s*registration\s*number|tax\s*id)[\s:â€“-]*([3]\d{13}[3])", txt, re.I)
            if not vat_m:
                vat_m = re.search(r"\b([3]\d{13}[3])\b", txt)
            if not vat_m:
                vat_m = re.search(r"\b(\d{15})\b", txt)
            if vat_m:
                self.tax_id = vat_m.group(1)

            # 3. GOSI / Unified 700 Number
            gosi_m = re.search(r"(?:Ø±Ù‚Ù…\s*Ø§Ù„Ø§Ø´ØªØ±Ø§Ùƒ|Ø±Ù‚Ù…\s*Ø§Ø´ØªØ±Ø§Ùƒ\s*Ø§Ù„Ù…Ù†Ø´Ø£Ø©|Ø§Ù„ØªØ£Ù…ÙŠÙ†Ø§Øª\s*Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ©|gosi\s*no)[\s:â€“-]*(\d{7,10})", txt, re.I)
            if gosi_m:
                self.gosi_number = gosi_m.group(1)
            else:
                u700_m = re.search(r"(?:Ø§Ù„Ø±Ù‚Ù…\s*Ø§Ù„ÙˆØ·Ù†ÙŠ\s*Ø§Ù„Ù…ÙˆØ­Ø¯|Ø§Ù„Ø±Ù‚Ù…\s*Ø§Ù„Ù…ÙˆØ­Ø¯)[\s:â€“-]*([7]\d{9})", txt, re.I)
                if u700_m:
                    self.gosi_number = u700_m.group(1)

            # 4. Balady License No
            balady_m = re.search(r"(?:Ø±Ø®ØµØ©\s*Ø¨Ù„Ø¯ÙŠØ©|Ø±Ø®ØµØ©\s*Ù†Ø´Ø§Ø·\s*ØªØ¬Ø§Ø±ÙŠ|Ù…Ù†ØµØ©\s*Ø¨Ù„Ø¯ÙŠ|Ø¨Ù„Ø¯ÙŠ|Ø±Ù‚Ù…\s*Ø§Ù„Ø±Ø®ØµØ©)[\s:â€“-]*(\d{8,14})", txt, re.I)
            if balady_m:
                self.balady_license_no = balady_m.group(1)

            # 5. National Address Details (Building No, Postal Code, Additional No, District, Street, Short Address)
            bm = re.search(r"(?:Ø±Ù‚Ù…\s*Ø§Ù„Ù…Ø¨Ù†Ù‰|building\s*no)[\s:â€“-]*(\d{4})", txt, re.I)
            if bm:
                self.building_no = bm.group(1)
            pm = re.search(r"(?:Ø§Ù„Ø±Ù…Ø²\s*Ø§Ù„Ø¨Ø±ÙŠØ¯ÙŠ|postal\s*code|zip)[\s:â€“-]*(\d{5})", txt, re.I)
            if pm:
                self.postal_code = pm.group(1)
            am = re.search(r"(?:Ø§Ù„Ø±Ù‚Ù…\s*Ø§Ù„Ø¥Ø¶Ø§ÙÙŠ|Ø§Ù„Ø±Ù‚Ù…\s*Ø§Ù„ÙØ±Ø¹ÙŠ|secondary\s*no|additional\s*no)[\s:â€“-]*(\d{4})", txt, re.I)
            if am:
                self.additional_no = am.group(1)
            dm = re.search(r"(?:Ø§Ù„Ø­ÙŠ|district|Ø­ÙŠ)[\s:â€“-]*([^\n,]+)", txt, re.I)
            if dm:
                self.district = dm.group(1).strip()
            sm = re.search(r"(?:Ø§Ù„Ø´Ø§Ø±Ø¹|street|Ø·Ø±ÙŠÙ‚)[\s:â€“-]*([^\n,]+)", txt, re.I)
            if sm:
                self.street = sm.group(1).strip()
            short_addr_m = re.search(r"(?:Ø§Ù„Ø¹Ù†ÙˆØ§Ù†\s*Ø§Ù„Ù…Ø®ØªØµØ±|short\s*address)[\s:â€“-]*([A-Za-z]{4}\d{4})", txt, re.I)
            if short_addr_m:
                self.national_short_address = short_addr_m.group(1).upper()

            # 6. Employee count & Saudization
            emp_m = re.search(r"(?:Ø§Ù„Ù…Ø¬Ù…ÙˆØ¹|Ø¥Ø¬Ù…Ø§Ù„ÙŠ\s*Ø§Ù„Ø¹Ø§Ù…Ù„ÙŠÙ†|Ø¹Ø¯Ø¯\s*Ø§Ù„Ù…Ø´ØªØ±ÙƒÙŠÙ†|total\s*employees)[\s:â€“-]*(\d+)", txt, re.I)
            if emp_m:
                self.employee_count = int(emp_m.group(1))

            saudi_emp_m = re.search(r"(?:Ø¹Ø¯Ø¯\s*Ø§Ù„Ù…Ø´ØªØ±ÙƒÙŠÙ†\s*Ø§Ù„Ø³Ø¹ÙˆØ¯ÙŠÙŠÙ†|Ø³Ø¹ÙˆØ¯ÙŠÙŠÙ†)[\s:â€“-]*(\d+)", txt, re.I)
            if saudi_emp_m and self.employee_count:
                saudi_cnt = int(saudi_emp_m.group(1))
                self.saudization_rate = round((saudi_cnt / self.employee_count) * 100.0, 1)

            # 7. Extract Exact Company Name
            for line in txt.splitlines():
                line = line.strip()
                if any(kw in line for kw in ["Ø´Ø±ÙƒØ©", "Ù…Ø¤Ø³Ø³Ø©", "ÙØ±Ø¹ Ø´Ø±ÙƒØ©", "Ù…Ø¬Ù…ÙˆØ¹Ø©", "Ù…ØµÙ†Ø¹"]):
                    clean = re.sub(r"(?:ØªØ´Ù‡Ø¯\s*Ø§Ù„ØºØ±ÙØ©\s*Ø¨Ø£Ù†|ØªØ´Ù‡Ø¯\s*Ø§Ù„Ù…Ø¤Ø³Ø³Ø©\s*Ø¨Ø£Ù†|Ø¨Ø£Ù†|Ø§Ø³Ù…\s*Ø§Ù„Ù…Ù†Ø´Ø£Ø©|Ø§Ø³Ù…\s*Ø§Ù„Ø´Ø±ÙƒØ©|Ø§Ø³Ù…\s*Ø§Ù„Ù…Ø¤Ø³Ø³Ø©|Ø§Ù„Ø§Ø³Ù…\s*Ø§Ù„ØªØ¬Ø§Ø±ÙŠ\s*Ù„Ù„Ø´Ø±ÙƒØ©|Ø§Ù„Ø§Ø³Ù…\s*Ø§Ù„ØªØ¬Ø§Ø±ÙŠ|Ø§Ø³Ù…\s*Ø§Ù„Ù…ÙƒÙ„Ù|Ø§Ù„Ø§Ø³Ù…)[\s:â€“-]*", "", line).strip()
                    # Clean punctuation and brackets
                    clean = re.sub(r"^[^\w\u0600-\u06FF]+|[^\w\u0600-\u06FF]+$", "", clean).strip()
                    if 5 < len(clean) < 100:
                        self.company_name = clean
                        break

            # 8. Industry Domain
            if any(w in full_lower for w in ["ØªÙ‚Ù†ÙŠØ©", "Ù…Ø¹Ù„ÙˆÙ…Ø§Øª", "Ø¨Ø±Ù…Ø¬ÙŠØ§Øª", "ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ù†Ø¸Ù…", "software", "it"]):
                self.industry_domain = "services"
            elif any(w in full_lower for w in ["Ù…Ø·Ø¹Ù…", "ÙƒØ§ÙÙŠÙ‡", "Ù…Ù‚Ù‡Ù‰", "Ø£ØºØ°ÙŠØ©", "ÙˆØ¬Ø¨Ø§Øª", "restaurant"]):
                self.industry_domain = "restaurant"
            elif any(w in full_lower for w in ["ØªØµÙ†ÙŠØ¹", "Ù…ØµÙ†Ø¹", "ØµÙ†Ø§Ø¹ÙŠ", "manufacturing"]):
                self.industry_domain = "manufacturing"
            elif any(w in full_lower for w in ["Ù…Ù‚Ø§ÙˆÙ„Ø§Øª", "Ø¨Ù†Ø§Ø¡", "ØªØ´ÙŠÙŠØ¯", "Ø¹Ù‚ÙˆØ¯", "construction"]):
                self.industry_domain = "construction"
            elif any(w in full_lower for w in ["Ù…Ø­Ø·Ø©", "ÙˆÙ‚ÙˆØ¯", "Ø¨Ù†Ø²ÙŠÙ†", "fuel"]):
                self.industry_domain = "fuel_station"
            elif any(w in full_lower for w in ["Ø¹Ù‚Ø§Ø±", "Ø¹Ù‚Ø§Ø±Ø§Øª", "real estate"]):
                self.industry_domain = "real_estate"

            for c in ["Ø§Ù„Ø±ÙŠØ§Ø¶", "Ø¬Ø¯Ø©", "Ø§Ù„Ø¯Ù…Ø§Ù…", "Ù…ÙƒØ©", "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©", "Ø§Ù„Ø®Ø¨Ø±", "Ø§Ù„Ù‚ØµÙŠÙ…", "ØªØ¨ÙˆÙƒ", "Ø£Ø¨Ù‡Ø§"]:
                if c in txt:
                    self.city = c
                    break

            # Classify
            doc_type = "other"
            doc_label = "ðŸ“„ ÙˆØ«ÙŠÙ‚Ø© Ø£Ø¹Ù…Ø§Ù„ Ø¹Ø§Ù…Ø©"
            summary_info = []

            if "Ø³Ø¬Ù„ ØªØ¬Ø§Ø±ÙŠ" in txt or "ÙˆØ²Ø§Ø±Ø© Ø§Ù„ØªØ¬Ø§Ø±Ø©" in txt or (cr_m and "Ø±Ø£Ø³ Ø§Ù„Ù…Ø§Ù„" in txt):
                doc_type = "cr"
                doc_label = "ðŸ“‘ Ø§Ù„Ø³Ø¬Ù„ Ø§Ù„ØªØ¬Ø§Ø±ÙŠ"
                if cr_m:
                    summary_info.append(f"Ø±Ù‚Ù… Ø§Ù„Ø³Ø¬Ù„: {cr_m.group(1)}")
            elif "Ø§Ù„Ø²ÙƒØ§Ø© ÙˆØ§Ù„Ø¶Ø±ÙŠØ¨Ø© ÙˆØ§Ù„Ø¬Ù…Ø§Ø±Ùƒ" in txt or "zatca" in full_lower or (vat_m and ("Ø¶Ø±ÙŠØ¨Ø© Ø§Ù„Ù‚ÙŠÙ…Ø© Ø§Ù„Ù…Ø¶Ø§ÙØ©" in txt or "Ø´Ù‡Ø§Ø¯Ø© ØªØ³Ø¬ÙŠÙ„" in txt)):
                doc_type = "vat"
                doc_label = "ðŸ§¾ Ø´Ù‡Ø§Ø¯Ø© Ø¶Ø±ÙŠØ¨Ø© Ø§Ù„Ù‚ÙŠÙ…Ø© Ø§Ù„Ù…Ø¶Ø§ÙØ©"
                if vat_m:
                    summary_info.append(f"Ø§Ù„Ø±Ù‚Ù… Ø§Ù„Ø¶Ø±ÙŠØ¨ÙŠ: {vat_m.group(1)}")
            elif "Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„ÙˆØ·Ù†ÙŠ" in txt or "national address" in full_lower or "Ø³Ø¨Ù„" in txt or (bm and pm):
                doc_type = "national_address"
                doc_label = "ðŸ“ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„ÙˆØ·Ù†ÙŠ"
                if bm and pm:
                    summary_info.append(f"Ù…Ø¨Ù†Ù‰ {bm.group(1)} - Ø¨Ø±ÙŠØ¯ {pm.group(1)}")
            elif "Ø§Ù„ØªØ£Ù…ÙŠÙ†Ø§Øª Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ©" in txt or "gosi" in full_lower:
                doc_type = "gosi"
                doc_label = "ðŸ›¡ï¸ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„ØªØ£Ù…ÙŠÙ†Ø§Øª Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ©"
                if gosi_m:
                    summary_info.append(f"Ø±Ù‚Ù… Ø§Ù„Ù…Ù†Ø´Ø£Ø©: {gosi_m.group(1)}")
            elif "Ø¨Ù„Ø¯ÙŠ" in txt or "Ø±Ø®ØµØ© Ø§Ù„Ù†Ø´Ø§Ø· Ø§Ù„ØªØ¬Ø§Ø±ÙŠ" in txt or balady_m:
                doc_type = "balady"
                doc_label = "ðŸ¢ Ø±Ø®ØµØ© Ø¨Ù„Ø¯ÙŠ"
                if balady_m:
                    summary_info.append(f"Ø±Ù‚Ù… Ø§Ù„Ø±Ø®ØµØ©: {balady_m.group(1)}")
            elif "Ù†Ø·Ø§Ù‚Ø§Øª" in txt or "Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„Ø³Ø¹ÙˆØ¯Ø©" in txt or "ÙˆØ²Ø§Ø±Ø© Ø§Ù„Ù…ÙˆØ§Ø±Ø¯ Ø§Ù„Ø¨Ø´Ø±ÙŠØ©" in txt:
                doc_type = "nitaqat"
                doc_label = "ðŸ‘¥ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„Ø³Ø¹ÙˆØ¯Ø© ÙˆÙ†Ø·Ø§Ù‚Ø§Øª"
            elif "Ø§Ù„ØºØ±ÙØ© Ø§Ù„ØªØ¬Ø§Ø±ÙŠØ©" in txt or "Ø§Ø´ØªØ±Ø§Ùƒ Ø§Ù„ØºØ±ÙØ©" in txt:
                doc_type = "chamber"
                doc_label = "ðŸ›ï¸ Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„ØºØ±ÙØ© Ø§Ù„ØªØ¬Ø§Ø±ÙŠØ©"

            # Create or update Scanned Document Record
            scanned_doc_obj.create({
                "journey_id": self.id,
                "name": item["name"],
                "attachment_id": item["attachment"].id if item["attachment"] else False,
                "document_type": doc_type,
                "extracted_summary": " | ".join(summary_info) if summary_info else doc_label,
                "state": "extracted",
            })

            detected_types_summary.append(doc_label)
            processed_count += 1

        # Build Interactive Dashboard of Extracted Data
        badges = []
        if self.cr_number:
            badges.append(f'<span class="badge bg-success p-2 me-2 mb-1" style="font-size:13px;">ðŸ“‘ Ø§Ù„Ø³Ø¬Ù„: {self.cr_number}</span>')
        if self.tax_id:
            badges.append(f'<span class="badge bg-success p-2 me-2 mb-1" style="font-size:13px;">ðŸ§¾ Ø§Ù„Ø¶Ø±ÙŠØ¨Ø©: {self.tax_id}</span>')
        if self.gosi_number:
            badges.append(f'<span class="badge bg-success p-2 me-2 mb-1" style="font-size:13px;">ðŸ›¡ï¸ Ø§Ù„ØªØ£Ù…ÙŠÙ†Ø§Øª: {self.gosi_number}</span>')
        if self.balady_license_no:
            badges.append(f'<span class="badge bg-success p-2 me-2 mb-1" style="font-size:13px;">ðŸ¢ Ø±Ø®ØµØ© Ø¨Ù„Ø¯ÙŠ: {self.balady_license_no}</span>')
        if self.building_no or self.district:
            addr_str = f"{self.city or ''} - {self.district or ''} - Ù…Ø¨Ù†Ù‰ {self.building_no or ''}"
            badges.append(f'<span class="badge bg-info p-2 me-2 mb-1" style="font-size:13px;">ðŸ“ Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„ÙˆØ·Ù†ÙŠ: {addr_str}</span>')
        if self.saudization_rate:
            badges.append(f'<span class="badge bg-primary p-2 me-2 mb-1" style="font-size:13px;">ðŸ‘¥ Ù†Ø³Ø¨Ø© Ø§Ù„ØªÙˆØ·ÙŠÙ†: {self.saudization_rate}%</span>')

        self.document_scan_summary_html = f"""
            <div class="alert alert-success border-0 shadow-sm p-3 mb-3" style="border-radius: 10px; background-color: #E8F5E9;">
                <div class="d-flex justify-content-between align-items-center mb-2">
                    <strong class="text-success" style="font-size: 16px;">
                        ðŸŽ¯ ØªÙ… Ø¨Ù†Ø¬Ø§Ø­ ÙØ­Øµ ÙˆØµÙŠØ¯ Ø¨ÙŠØ§Ù†Ø§Øª ({processed_count}) ÙˆØ«ÙŠÙ‚Ø© Ø¯ÙØ¹Ø© ÙˆØ§Ø­Ø¯Ø©!
                    </strong>
                    <span class="badge bg-success py-1 px-2">{processed_count} Ù…Ù„ÙØ§Øª Ù…Ø¹Ø§Ù„Ø¬Ø©</span>
                </div>
                <p class="text-dark small mb-2">Ù‚Ø§Ù… Ø§Ù„Ø°ÙƒØ§Ø¡ Ø§Ù„Ø§ØµØ·Ù†Ø§Ø¹ÙŠ Ø¨ØªØµÙ†ÙŠÙ Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ù„ÙØ§Øª ÙˆØªØ­Ø¯ÙŠØ« Ø³Ø¬Ù„ Ø§Ù„Ù…Ù†Ø´Ø£Ø©:</p>
                <div class="d-flex flex-wrap mt-1">
                    {' '.join(badges)}
                </div>
            </div>
        """

        # Sync to Company Record directly
        if self.company_id:
            comp_vals = {}
            if self.company_name:
                comp_vals["name"] = self.company_name
            if self.tax_id:
                comp_vals["vat"] = self.tax_id
            if self.city:
                comp_vals["city"] = self.city
            if self.street:
                comp_vals["street"] = self.street
            if self.district:
                comp_vals["street2"] = self.district
            if self.postal_code:
                comp_vals["zip"] = self.postal_code
            if comp_vals:
                self.company_id.write(comp_vals)

        # Reset single dropzone
        self.upload_document_file = False
        self.upload_document_filename = False

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("ðŸŽ¯ ØªÙ… ÙØ­Øµ ÙˆØµÙŠØ¯ Ø§Ù„Ø¯ÙØ¹Ø© Ø¨Ù†Ø¬Ø§Ø­!"),
                "message": _("ØªÙ…Øª Ù…Ø¹Ø§Ù„Ø¬Ø© ÙˆØªØµÙ†ÙŠÙ %d ÙˆØ«ÙŠÙ‚Ø© Ø±Ø³Ù…ÙŠØ© ÙˆØ§Ø³ØªØ®Ø±Ø§Ø¬ ÙƒØ§ÙØ© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø¨Ù†Ø¬Ø§Ø­.") % processed_count,
                "type": "success",
                "sticky": False,
            },
        }

    def action_open_universal_migrator(self):
        """Open the Universal Enterprise Migrator for SAP / Oracle / Microsoft / SQL / Excel."""
        return {
            "name": "ðŸ“¦ Ø§Ù„Ù…Ø³ØªÙˆØ±Ø¯ ÙˆØ§Ù„Ù…Ù‡Ø¬Ø± Ø§Ù„Ø°ÙƒÙŠ Ø§Ù„Ø´Ø§Ù…Ù„ (SAP / Oracle / Microsoft / SQL)",
            "type": "ir.actions.act_window",
            "res_model": "nexus.universal.migrator",
            "view_mode": "form",
            "target": "new",
        }

    def action_clear_hr_inputs(self):
        """Clear uploaded employee documents and roster sheets."""
        self.ensure_one()
        self.employee_document_attachment_ids = [(5, 0, 0)]
        self.hr_roster_file = False
        self.hr_roster_filename = False
        self.hr_scan_summary_html = False
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("ðŸ§¹ ØªÙ… Ù…Ø³Ø­ Ù…Ø¯Ø®Ù„Ø§Øª Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†"),
                "message": _("ØªÙ… ØªÙØ±ÙŠØº Ù…Ø±ÙÙ‚Ø§Øª Ø§Ù„Ù‡ÙˆÙŠØ§Øª ÙˆÙƒØ´ÙˆÙ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø¨Ù†Ø¬Ø§Ø­."),
                "type": "info",
                "sticky": False,
            },
        }

    def action_scan_employee_documents(self):
        """Scan uploaded Iqamas & Passports and auto-create HR Employee records."""
        self.ensure_one()
        import base64
        import re

        attachments = self.employee_document_attachment_ids
        if not attachments:
            raise UserError(_("ÙŠØ±Ø¬Ù‰ Ø³Ø­Ø¨ Ø£Ùˆ Ø±ÙØ¹ ØµÙˆØ± Ø£Ùˆ Ù…Ù„ÙØ§Øª PDF Ù„Ø¥Ù‚Ø§Ù…Ø§Øª ÙˆØ¬ÙˆØ§Ø²Ø§Øª Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø£ÙˆÙ„Ø§Ù‹."))

        Employee = self.env["hr.employee"]
        Job = self.env["hr.job"]
        created_employees = []
        created_badges = []

        for att in attachments:
            if not att.datas:
                continue
            try:
                raw_bytes = base64.b64decode(att.datas)
                txt = _normalize_arabic_digits(raw_bytes.decode("utf-8", errors="ignore"))
            except Exception:
                continue

            # 1. Iqama / National ID (10 digits starting with 1 or 2)
            iq_m = re.search(r"(?:Ø±Ù‚Ù…\s*Ø§Ù„Ø¥Ù‚Ø§Ù…Ø©|Ø±Ù‚Ù…\s*Ø§Ù„Ù‡ÙˆÙŠØ©|iqama\s*no|id\s*no)[\s:â€“-]*([12]\d{9})", txt, re.I)
            if not iq_m:
                iq_m = re.search(r"\b([12]\d{9})\b", txt)
            iqama_no = iq_m.group(1) if iq_m else False

            # 2. Passport Number & MRZ
            pass_m = re.search(r"(?:passport\s*no|Ø±Ù‚Ù…\s*Ø§Ù„Ø¬ÙˆØ§Ø²)[\s:â€“-]*([A-Za-z0-9]{7,10})", txt, re.I)
            passport_no = pass_m.group(1).upper() if pass_m else False

            # 3. Full Name
            emp_name = None
            name_m = re.search(r"(?:Ø§Ù„Ø§Ø³Ù…|Ø§Ø³Ù…\s*Ø§Ù„Ù…Ù‚ÙŠÙ…|Ø§Ø³Ù…\s*Ø§Ù„Ù…ÙˆØ§Ø·Ù†|name|full\s*name)[\s:â€“-]*([^\n]+)", txt, re.I)
            if name_m:
                emp_name = name_m.group(1).strip()
            else:
                for line in txt.splitlines():
                    clean_l = line.strip()
                    if 3 < len(clean_l.split()) <= 6 and not any(kw in clean_l for kw in ["Ø§Ù„Ù…Ù…Ù„ÙƒØ©", "ÙˆØ²Ø§Ø±Ø©", "Ø§Ù„Ø¥Ù‚Ø§Ù…Ø©", "Ø§Ù„Ø¬ÙˆØ§Ø²Ø§Øª", "ØªØ§Ø±ÙŠØ®"]):
                        emp_name = clean_l
                        break
            if not emp_name:
                emp_name = f"Ù…ÙˆØ¸Ù Ø¬Ø¯ÙŠØ¯ ({att.name})"

            # 4. Job Position / Profession (Ø§Ù„Ù…Ù‡Ù†Ø©)
            job_title = None
            job_m = re.search(r"(?:Ø§Ù„Ù…Ù‡Ù†Ø©|Ø§Ù„Ù…Ø³Ù…Ù‰\s*Ø§Ù„ÙˆØ¸ÙŠÙÙŠ|Ø§Ù„ÙˆØ¸ÙŠÙØ©|occupation|job\s*title)[\s:â€“-]*([^\n,]+)", txt, re.I)
            if job_m:
                job_title = job_m.group(1).strip()

            # Job record
            job_id = False
            if job_title:
                existing_job = Job.search([("name", "=", job_title), ("company_id", "=", self.company_id.id)], limit=1)
                if not existing_job:
                    existing_job = Job.create({"name": job_title, "company_id": self.company_id.id})
                job_id = existing_job.id

            # Create or update Employee
            emp_vals = {
                "name": emp_name,
                "company_id": self.company_id.id,
            }
            if iqama_no:
                emp_vals["identification_id"] = iqama_no
            if passport_no:
                emp_vals["passport_id"] = passport_no
            if job_id:
                emp_vals["job_id"] = job_id

            # Search existing by ID or passport
            existing_emp = False
            if iqama_no:
                existing_emp = Employee.search([("identification_id", "=", iqama_no), ("company_id", "=", self.company_id.id)], limit=1)
            if not existing_emp and passport_no:
                existing_emp = Employee.search([("passport_id", "=", passport_no), ("company_id", "=", self.company_id.id)], limit=1)

            if existing_emp:
                existing_emp.write(emp_vals)
                emp_rec = existing_emp
            else:
                emp_rec = Employee.create(emp_vals)

            created_employees.append(emp_rec)
            created_badges.append(
                f'<span class="badge bg-success p-2 me-2 mb-2" style="font-size:13px;">'
                f'ðŸ‘¤ {emp_rec.name} | Ù‡ÙˆÙŠØ©: {iqama_no or passport_no or "Ù…Ø¹ØªÙ…Ø¯"} | ÙˆØ¸ÙŠÙØ©: {job_title or "Ø¹Ø§Ù…"}'
                f'</span>'
            )

        # Link to journey employee_ids
        if created_employees:
            self.employee_ids = [(4, emp.id) for emp in created_employees]
            self.employee_count = len(self.employee_ids)

        count = len(created_employees)
        self.hr_scan_summary_html = f"""
            <div class="alert alert-success border-0 shadow-sm p-3 mb-3" style="border-radius: 10px; background-color: #E8F5E9;">
                <h6 class="text-success font-weight-bold mb-2">ðŸŽ¯ ØªÙ… ØµÙŠØ¯ ÙˆØªØ£Ø³ÙŠØ³ ({count}) Ù…ÙˆØ¸ÙØ§Ù‹ Ù…Ù† ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ø¥Ù‚Ø§Ù…Ø§Øª ÙˆØ§Ù„Ø¬ÙˆØ§Ø²Ø§Øª Ø¨Ù†Ø¬Ø§Ø­:</h6>
                <div class="d-flex flex-wrap mt-2">
                    {' '.join(created_badges)}
                </div>
            </div>
        """

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("ðŸŽ¯ ØªÙ… Ø§Ø³ØªØ®Ø±Ø§Ø¬ ÙˆØªØ£Ø³ÙŠØ³ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø¨Ù†Ø¬Ø§Ø­!"),
                "message": _("ØªÙ…Øª Ù…Ø¹Ø§Ù„Ø¬Ø© ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ù‡ÙˆÙŠØ§Øª ÙˆØªØ£Ø³ÙŠØ³ %d Ù…ÙˆØ¸ÙØ§Ù‹ ÙÙŠ Ø§Ù„Ù†Ø¸Ø§Ù….") % count,
                "type": "success",
                "sticky": False,
            },
        }

    def action_import_employee_roster(self):
        """Smart Roster Importer â€” Parses Excel (.xlsx), CSV, or text roster and bulk-provisions employees."""
        self.ensure_one()
        import base64
        import csv
        import io
        import re

        if not self.hr_roster_file:
            raise UserError(_("ÙŠØ±Ø¬Ù‰ Ø¥Ø±ÙØ§Ù‚ Ù…Ù„Ù ÙƒØ´Ù Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† (Excel .xlsx Ø£Ùˆ CSV Ø£Ùˆ PDF) Ø£ÙˆÙ„Ø§Ù‹."))

        try:
            raw_bytes = base64.b64decode(self.hr_roster_file)
        except Exception:
            raise UserError(_("ØªØ¹Ø°Ø± Ù‚Ø±Ø§Ø¡Ø© Ù…Ù„Ù ÙƒØ´Ù Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†."))

        Employee = self.env["hr.employee"]
        Department = self.env["hr.department"]
        Job = self.env["hr.job"]
        Contract = self.env.get("hr.contract")

        imported_rows = []
        filename = (self.hr_roster_filename or "").lower()

        # Try parsing as Excel (.xlsx) using openpyxl or xlrd if available
        parsed_as_excel = False
        if filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        row_dict = {headers[idx]: (str(val).strip() if val is not None else '') for idx, val in enumerate(row) if idx < len(headers)}
                        imported_rows.append(row_dict)
                parsed_as_excel = True
            except Exception:
                pass

        # Fallback: Parse as CSV / Text with multi-delimiters
        if not parsed_as_excel:
            try:
                text_content = _normalize_arabic_digits(raw_bytes.decode("utf-8", errors="ignore"))
                # Detect delimiter
                first_line = text_content.splitlines()[0] if text_content.splitlines() else ""
                delim = ","
                for d in ["\t", ";", ",", "|"]:
                    if d in first_line:
                        delim = d
                        break
                reader = csv.DictReader(io.StringIO(text_content), delimiter=delim)
                for row in reader:
                    imported_rows.append({k.strip(): str(v).strip() for k, v in row.items() if k})
            except Exception:
                pass

        if not imported_rows:
            raise UserError(_("ØªØ¹Ø°Ø± Ù‚Ø±Ø§Ø¡Ø© Ø£Ø³Ø·Ø± Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ù…Ù† Ø§Ù„Ù…Ù„Ù. ÙŠØ±Ø¬Ù‰ Ø§Ù„ØªØ£ÙƒØ¯ Ù…Ù† Ø£Ù† Ø§Ù„Ù…Ù„Ù Ø¨ØµÙŠØºØ© Excel Ø£Ùˆ CSV ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰ Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ø£Ø³Ù…Ø§Ø¡ ÙˆØ§Ù„Ù‡ÙˆÙŠØ§Øª."))

        created_emps = []
        created_badges = []

        for row in imported_rows:
            # Match columns by Arabic/English keywords
            name = None
            iqama = None
            job = None
            dept = None
            salary = 0.0

            for k, val in row.items():
                k_lower = k.lower()
                val_clean = _normalize_arabic_digits(val)
                if any(w in k_lower for w in ["Ø§Ù„Ø§Ø³Ù…", "Ø§Ø³Ù… Ø§Ù„Ù…ÙˆØ¸Ù", "name", "employee name"]):
                    name = val
                elif any(w in k_lower for w in ["Ø§Ù„Ø¥Ù‚Ø§Ù…Ø©", "Ø§Ù„Ù‡ÙˆÙŠØ©", "iqama", "id", "national_id"]):
                    m = re.search(r"([12]\d{9})", val_clean)
                    iqama = m.group(1) if m else val_clean
                elif any(w in k_lower for w in ["Ø§Ù„Ù…Ù‡Ù†Ø©", "Ø§Ù„ÙˆØ¸ÙŠÙØ©", "Ø§Ù„Ù…Ø³Ù…Ù‰", "job", "position", "title"]):
                    job = val
                elif any(w in k_lower for w in ["Ø§Ù„Ù‚Ø³Ù…", "Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©", "department", "dept"]):
                    dept = val
                elif any(w in k_lower for w in ["Ø§Ù„Ø±Ø§ØªØ¨", "Ø§Ù„Ø£Ø³Ø§Ø³ÙŠ", "salary", "wage", "basic"]):
                    try:
                        salary = float(re.sub(r"[^\d.]", "", val_clean))
                    except Exception:
                        pass

            if not name:
                continue

            # Resolve Department
            dept_id = False
            if dept:
                existing_dept = Department.search([("name", "=", dept), ("company_id", "=", self.company_id.id)], limit=1)
                if not existing_dept:
                    existing_dept = Department.create({"name": dept, "company_id": self.company_id.id})
                dept_id = existing_dept.id
                if existing_dept.id not in self.department_ids.ids:
                    self.department_ids = [(4, existing_dept.id)]

            # Resolve Job
            job_id = False
            if job:
                existing_job = Job.search([("name", "=", job), ("company_id", "=", self.company_id.id)], limit=1)
                if not existing_job:
                    existing_job = Job.create({"name": job, "company_id": self.company_id.id})
                job_id = existing_job.id

            # Create/Update Employee
            emp_vals = {
                "name": name,
                "company_id": self.company_id.id,
            }
            if iqama:
                emp_vals["identification_id"] = iqama
            if job_id:
                emp_vals["job_id"] = job_id
            if dept_id:
                emp_vals["department_id"] = dept_id

            existing_emp = False
            if iqama:
                existing_emp = Employee.search([("identification_id", "=", iqama), ("company_id", "=", self.company_id.id)], limit=1)
            if not existing_emp:
                existing_emp = Employee.search([("name", "=", name), ("company_id", "=", self.company_id.id)], limit=1)

            if existing_emp:
                existing_emp.write(emp_vals)
                emp_rec = existing_emp
            else:
                emp_rec = Employee.create(emp_vals)

            # Auto-create Wage Contract for WPS compliance if salary present
            if Contract and salary > 0:
                existing_contract = Contract.search([("employee_id", "=", emp_rec.id), ("state", "=", "open")], limit=1)
                contract_vals = {
                    "name": f"Ø¹Ù‚Ø¯ Ø¹Ù…Ù„ - {emp_rec.name}",
                    "employee_id": emp_rec.id,
                    "company_id": self.company_id.id,
                    "wage": salary,
                    "state": "open",
                }
                if not existing_contract:
                    Contract.create(contract_vals)
                else:
                    existing_contract.write({"wage": salary})

            created_emps.append(emp_rec)
            created_badges.append(
                f'<span class="badge bg-success p-2 me-2 mb-2" style="font-size:13px;">'
                f'ðŸ‘¤ {emp_rec.name} | Ù‡ÙˆÙŠØ©: {iqama or "Ù…Ø¹ØªÙ…Ø¯"} | Ù‚Ø³Ù…: {dept or "Ø¹Ø§Ù…"} | Ø±Ø§ØªØ¨: {salary:,.0f} Ø±ÙŠØ§Ù„'
                f'</span>'
            )

        if created_emps:
            self.employee_ids = [(4, emp.id) for emp in created_emps]
            self.employee_count = len(self.employee_ids)

        count = len(created_emps)
        self.hr_scan_summary_html = f"""
            <div class="alert alert-success border-0 shadow-sm p-3 mb-3" style="border-radius: 10px; background-color: #E8F5E9;">
                <h6 class="text-success font-weight-bold mb-2">ðŸ“Š ØªÙ… Ø§Ø³ØªÙŠØ±Ø§Ø¯ ÙˆØªØ£Ø³ÙŠØ³ ({count}) Ù…ÙˆØ¸ÙØ§Ù‹ Ù…Ø¹ Ø¹Ù‚ÙˆØ¯ Ø§Ù„Ø±ÙˆØ§ØªØ¨ (WPS) Ø¨Ù†Ø¬Ø§Ø­:</h6>
                <div class="d-flex flex-wrap mt-2">
                    {' '.join(created_badges)}
                </div>
            </div>
        """

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("ðŸ“Š ØªÙ… Ø§Ø³ØªÙŠØ±Ø§Ø¯ ÙƒØ´Ù Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø¨Ù†Ø¬Ø§Ø­!"),
                "message": _("ØªÙ…Øª Ù…Ø¹Ø§Ù„Ø¬Ø© Ø§Ù„ÙƒØ´Ù ÙˆØªØ£Ø³ÙŠØ³ %d Ù…ÙˆØ¸ÙØ§Ù‹ Ù…Ø¹ Ù…Ø³ÙŠØ±Ø§Øª Ø§Ù„Ø±ÙˆØ§ØªØ¨ Ù„Ø­Ù…Ø§ÙŠØ© Ø§Ù„Ø£Ø¬ÙˆØ±.") % count,
                "type": "success",
                "sticky": False,
            },
        }

    def action_go_to_next_stage(self):
        """Move smoothly to the next setup stage."""
        self.ensure_one()
        stages = ["identity_legal", "structure_hr", "financial_core", "operations", "go_live", "done"]
        idx = stages.index(self.stage) if self.stage in stages else 0
        if idx < len(stages) - 1:
            self.stage = stages[idx + 1]
            if self.state == 'draft':
                self.state = 'in_progress'
        return True

    def action_go_to_prev_stage(self):
        """Move back to the previous setup stage."""
        self.ensure_one()
        stages = ["identity_legal", "structure_hr", "financial_core", "operations", "go_live", "done"]
        idx = stages.index(self.stage) if self.stage in stages else 0
        if idx > 0:
            self.stage = stages[idx - 1]
        return True
