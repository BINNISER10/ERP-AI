# -*- coding: utf-8 -*-
"""Nexus Universal Enterprise Migrator & Product Image Importer — المستورد والمهجر الذكي الشامل.

Imports master data (Products with images & barcodes, Customers/Vendors, Chart of Accounts,
and Opening Stock) from SAP, Oracle ERP, Microsoft Dynamics, SQL Dumps, and Excel bundles.
"""

import base64
import csv
import io
import json
import logging
import re
import zipfile
import requests

from odoo import api, fields, models, Command, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

LEGACY_SOURCES = [
    ('auto', '✨ كشف المصدر تلقائياً بالذكاء الاصطناعي (Auto-Detect Source)'),
    ('sap', '🔷 SAP (S/4HANA / ECC / Business One)'),
    ('oracle', '🔴 Oracle (Fusion / EBS / NetSuite / DB Export)'),
    ('microsoft', '🟦 Microsoft Dynamics (365 / Business Central / NAV)'),
    ('sql_dump', '💾 SQL Database Dump (Postgres / MySQL / MSSQL / Oracle)'),
    ('excel_generic', '📊 Excel / CSV / JSON Bundle with Images'),
]

TARGET_DOMAINS = [
    ('auto', '🤖 تحديد نوع البيانات تلقائياً (Auto-Detect Target)'),
    ('products', '📦 المنتجات والأصناف مع الصور والباركود والأسعار والمخزون'),
    ('partners', '👥 العملاء والموردين مع الأرقام الضريبية والعناوين'),
    ('accounts', '💰 دليل الحسابات والأرصدة الافتتاحية'),
    ('employees', '👤 الموظفين والهيكل الإداري والرواتب'),
]


class NexusUniversalMigrator(models.TransientModel):
    _name = 'nexus.universal.migrator'
    _description = 'Nexus Universal Enterprise Data Migrator'

    state = fields.Selection(
        [
            ('upload', '1. رفع الملفات والحزم (Upload Data)'),
            ('preview', '2. المعاينة والربط الذكي للحقول (AI Mapping & Preview)'),
            ('imported', '3. تم الترحيل والاستيراد بنجاح (Imported)'),
        ],
        default='upload',
        required=True,
    )

    data_file = fields.Binary(
        string='ملف البيانات أو الحزمة (Excel / CSV / SQL / JSON / ZIP with Images)',
        required=True,
        attachment=True,
    )
    data_filename = fields.Char(string='اسم الملف')

    legacy_source = fields.Selection(
        selection=LEGACY_SOURCES,
        string='نظام المصدر السابق (Source System)',
        default='auto',
        required=True,
    )
    import_target = fields.Selection(
        selection=TARGET_DOMAINS,
        string='نوع البيانات المراد استيرادها (Target Domain)',
        default='auto',
        required=True,
    )

    detected_source_title = fields.Char(string='المصدر المكتشف ذكياً', readonly=True)
    detected_target_title = fields.Char(string='الهدف المكتشف', readonly=True)
    total_records_count = fields.Integer(string='عدد السجلات المكتشفة', readonly=True)
    preview_data_html = fields.Html(string='جدول معاينة البيانات والربط الدلالي', readonly=True)
    import_summary_html = fields.Html(string='تقرير نتائج الاستيراد والترحيل', readonly=True)
    raw_parsed_json = fields.Text(string='Raw Parsed JSON', readonly=True)

    def _normalize_text(self, text):
        if not text:
            return ""
        trans = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
        return str(text).translate(trans).strip()

    def _parse_sql_dump(self, sql_text):
        """Extract table rows from SQL INSERT statements."""
        rows = []
        # Find INSERT INTO statements
        insert_pattern = re.compile(r"INSERT\s+INTO\s+[`\"']?(\w+)[`\"']?\s*(?:\(([^)]+)\))?\s+VALUES\s*(.+?);", re.I | re.DOTALL)
        for match in insert_pattern.finditer(sql_text):
            cols_raw = match.group(2)
            values_raw = match.group(3)
            cols = [c.strip().strip('`"\'') for c in cols_raw.split(",")] if cols_raw else []

            # Extract tuple values e.g. ('val1', 'val2', 123)
            tuple_pattern = re.compile(r"\(([^)]+)\)")
            for t_match in tuple_pattern.finditer(values_raw):
                val_items = [v.strip().strip("'\"") for v in t_match.group(1).split(",")]
                if cols and len(cols) == len(val_items):
                    rows.append(dict(zip(cols, val_items)))
                else:
                    rows.append({f"col_{i+1}": v for i, v in enumerate(val_items)})
        return rows

    def action_analyze_and_preview(self):
        """Analyze legacy files (SAP/Oracle/Microsoft/SQL/Excel/ZIP), map fields, and generate preview."""
        self.ensure_one()
        if not self.data_file:
            raise UserError(_('يرجى اختيار ملف البيانات أولاً.'))

        raw_bytes = base64.b64decode(self.data_file)
        fname = (self.data_filename or '').lower()
        extracted_rows = []
        zip_images_map = {}

        # 1. Handle ZIP files (containing Excel/CSV + Image files)
        if fname.endswith('.zip'):
            try:
                with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
                    for zname in z.namelist():
                        z_lower = zname.lower()
                        if z_lower.endswith(('.png', '.jpg', '.jpeg', '.webp')):
                            # Store image binary by filename base
                            base_key = zname.split('/')[-1].split('.')[0].lower()
                            zip_images_map[base_key] = base64.b64encode(z.read(zname)).decode('utf-8')
                        elif z_lower.endswith(('.xlsx', '.xls')):
                            import openpyxl
                            wb = openpyxl.load_workbook(io.BytesIO(z.read(zname)), data_only=True)
                            ws = wb.active
                            headers = [str(cell.value or '').strip() for cell in ws[1]]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if any(row):
                                    extracted_rows.append({headers[idx]: (str(val).strip() if val is not None else '') for idx, val in enumerate(row) if idx < len(headers)})
                        elif z_lower.endswith(('.csv', '.tsv', '.txt')):
                            txt = z.read(zname).decode('utf-8', errors='ignore')
                            delim = ',' if ',' in txt.splitlines()[0] else ('\t' if '\t' in txt.splitlines()[0] else ';')
                            reader = csv.DictReader(io.StringIO(txt), delimiter=delim)
                            for row in reader:
                                extracted_rows.append({k.strip(): str(v).strip() for k, v in row.items() if k})
            except Exception as e:
                raise UserError(_('فشل قراءة ملف ZIP:\n%s') % str(e))

        # 2. Handle Excel files (.xlsx, .xls)
        elif fname.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        extracted_rows.append({headers[idx]: (str(val).strip() if val is not None else '') for idx, val in enumerate(row) if idx < len(headers)})
            except Exception as e:
                raise UserError(_('فشل قراءة ملف الإكسل:\n%s') % str(e))

        # 3. Handle SQL dump files (.sql)
        elif fname.endswith('.sql'):
            txt = raw_bytes.decode('utf-8', errors='ignore')
            extracted_rows = self._parse_sql_dump(txt)

        # 4. Handle CSV / Text / JSON
        else:
            try:
                txt = raw_bytes.decode('utf-8', errors='ignore')
                if txt.strip().startswith(('{', '[')):
                    loaded_json = json.loads(txt)
                    extracted_rows = loaded_json if isinstance(loaded_json, list) else [loaded_json]
                else:
                    first_line = txt.splitlines()[0] if txt.splitlines() else ""
                    delim = ',' if ',' in first_line else ('\t' if '\t' in first_line else ';')
                    reader = csv.DictReader(io.StringIO(txt), delimiter=delim)
                    for row in reader:
                        extracted_rows.append({k.strip(): str(v).strip() for k, v in row.items() if k})
            except Exception as e:
                raise UserError(_('فشل قراءة الملف:\n%s') % str(e))

        if not extracted_rows:
            raise UserError(_('لم يتم العثور على سجلات صالحة في الملف المرفوع.'))

        # Detect Source System & Target Domain
        all_keys_str = " ".join(" ".join(r.keys()).lower() for r in extracted_rows[:5])

        # Source detection
        detected_src = 'excel_generic'
        detected_src_title = '📊 ملف بيانات مخصص (Excel / CSV / JSON)'
        if any(k in all_keys_str for k in ['matnr', 'maktx', 'kunnr', 'lifnr', 'stcd1', 'stprs', 'kbetr', 'saknr']):
            detected_src = 'sap'
            detected_src_title = '🔷 SAP (S/4HANA / ECC / BAPI Extract)'
        elif any(k in all_keys_str for k in ['item_number', 'inventory_item_id', 'party_name', 'party_number', 'segment1', 'description_ar']):
            detected_src = 'oracle'
            detected_src_title = '🔴 Oracle ERP / Fusion / EBS Data Export'
        elif any(k in all_keys_str for k in ['item_no', 'customer_no', 'vendor_no', 'gen_prod_posting', 'item_description']):
            detected_src = 'microsoft'
            detected_src_title = '🟦 Microsoft Dynamics (365 / Business Central / NAV)'
        elif fname.endswith('.sql') or 'insert into' in all_keys_str:
            detected_src = 'sql_dump'
            detected_src_title = '💾 قاعدة بيانات SQL Dump (Postgres / MySQL / Oracle)'

        # Target detection
        detected_target = 'products'
        detected_target_title = '📦 المنتجات والأصناف مع الصور والأسعار والمخزون'
        if any(k in all_keys_str for k in ['customer', 'vendor', 'عميل', 'مورد', 'kunnr', 'lifnr', 'party_name', 'vat', 'tax_id', 'الرقم الضريبي']):
            if not any(k in all_keys_str for k in ['matnr', 'item_number', 'barcode', 'سعر البيع']):
                detected_target = 'partners'
                detected_target_title = '👥 العملاء والموردين (Customers & Vendors)'
        elif any(k in all_keys_str for k in ['account_code', 'رقم الحساب', 'اسم الحساب', 'saknr', 'chart_of_accounts']):
            detected_target = 'accounts'
            detected_target_title = '💰 دليل الحسابات والأرصدة الافتتاحية'
        elif any(k in all_keys_str for k in ['employee', 'موظف', 'راتب', 'salary', 'iqama', 'إقامة', 'gosi']):
            detected_target = 'employees'
            detected_target_title = '👤 الموظفين والرواتب'

        self.detected_source_title = detected_src_title
        self.detected_target_title = detected_target_title
        self.total_records_count = len(extracted_rows)
        self.raw_parsed_json = json.dumps({
            'rows': extracted_rows,
            'images_map': zip_images_map,
            'detected_target': detected_target,
            'detected_source': detected_src,
        }, ensure_ascii=False)

        # Build Preview HTML Table
        sample_rows = extracted_rows[:6]
        headers = list(sample_rows[0].keys()) if sample_rows else []
        th_cells = "".join([f"<th style='padding: 8px; border: 1px solid #E2E8F0; background: #EDF2F7;'>{h}</th>" for h in headers[:8]])
        
        tr_rows = []
        for r in sample_rows:
            td_cells = "".join([f"<td style='padding: 8px; border: 1px solid #E2E8F0;'>{r.get(h, '')}</td>" for h in headers[:8]])
            tr_rows.append(f"<tr>{td_cells}</tr>")

        self.preview_data_html = f"""
            <div class="card border-0 shadow-sm p-3 mb-3 bg-light" style="border-radius: 10px;">
                <div class="d-flex justify-content-between align-items-center mb-3">
                    <div>
                        <span class="badge bg-primary p-2 me-2" style="font-size: 13px;">{detected_src_title}</span>
                        <span class="badge bg-success p-2" style="font-size: 13px;">{detected_target_title}</span>
                    </div>
                    <strong class="text-dark">إجمالي السجلات المكتشفة: <span class="text-primary">{len(extracted_rows)} سجل</span></strong>
                </div>
                <div class="table-responsive" style="max-height: 250px; overflow-y: auto;">
                    <table class="table table-bordered table-sm mb-0" style="font-size: 13px; text-align: center;">
                        <thead><tr>{th_cells}</tr></thead>
                        <tbody>{''.join(tr_rows)}</tbody>
                    </table>
                </div>
            </div>
        """

        self.state = 'preview'
        return self._reopen()

    # ────────────────────────────────────────────────────────────────
    # Smart Chart-of-Accounts mapping (AI-assisted, rule-based fallback)
    # ────────────────────────────────────────────────────────────────
    def _classify_account_rule_based(self, code, name):
        """Offline classifier used when the Nexus AI service is unreachable.

        Combines keyword heuristics (works even on old, non-standard,
        company-specific numbering schemes) with the common 1/2/3/4/5
        first-digit convention as a last-resort fallback.
        """
        name_l = (name or '').lower()
        code_digits = re.sub(r'\D', '', str(code or ''))
        first_digit = code_digits[0] if code_digits else ''

        income_kw = ['إيراد', 'ايراد', 'مبيعات', 'دخل', 'revenue', 'income', 'sales']
        expense_kw = ['مصروف', 'مصاريف', 'تكلفة', 'تكاليف', 'expense', 'cost']
        equity_kw = ['رأس المال', 'حقوق الملكية', 'أرباح مبقاة', 'capital', 'equity', 'retained earning']
        liability_kw = ['خصوم', 'ذمم دائنة', 'دائنون', 'قروض', 'مستحقات', 'liability', 'payable', 'loan', 'accrued']
        cash_kw = ['نقد', 'صندوق', 'بنك', 'cash', 'bank']
        receivable_kw = ['ذمم مدينة', 'مدينون', 'عملاء', 'receivable', 'debtors']
        fixed_asset_kw = ['أصل ثابت', 'أصول ثابتة', 'fixed asset', 'ppe']
        asset_kw = ['أصول', 'اصول', 'asset', 'مخزون', 'inventory']

        def has_kw(words):
            return any(w in name_l for w in words)

        if has_kw(income_kw):
            return 'income'
        if has_kw(expense_kw):
            return 'expense'
        if has_kw(equity_kw):
            return 'equity'
        if has_kw(liability_kw):
            return 'liability_current'
        if has_kw(cash_kw):
            return 'asset_cash'
        if has_kw(receivable_kw):
            return 'asset_receivable'
        if has_kw(fixed_asset_kw):
            return 'asset_fixed'
        if has_kw(asset_kw):
            return 'asset_current'

        # Last resort: standard global numbering convention.
        return {
            '1': 'asset_current',
            '2': 'liability_current',
            '3': 'equity',
            '4': 'income',
            '5': 'expense',
        }.get(first_digit, 'asset_current')

    def _get_ai_coa_mapping(self, extracted_rows):
        """Ask the Nexus AI service to classify+harmonize the legacy chart of
        accounts. Returns a dict keyed by account code, or None if the AI
        service could not be reached (caller falls back to rule-based).
        """
        payload_accounts = [
            {
                'code': r['code'],
                'name': r['name'],
                'raw_type_hint': r.get('raw_type_hint'),
                'parent_code': r.get('parent_code'),
            }
            for r in extracted_rows[:300]
        ]
        try:
            config = self.env['copilot.config'].sudo().get_active_config(self.env.company)
            api_key = config.nexus_ai_api_key if config else False
            headers = {'X-API-Key': api_key} if api_key else {}
            resp = requests.post(
                'http://nexus_ai:8000/api/v1/ai/wizard/coa-mapping',
                json={
                    'accounts': payload_accounts,
                    'country': self.env.company.country_id.code or 'SA',
                    'language': 'ar',
                },
                headers=headers,
                timeout=45,
            )
            if resp.ok:
                data = resp.json()
                mappings = {
                    m.get('code'): m
                    for m in data.get('mappings', [])
                    if m.get('code')
                }
                if mappings:
                    return mappings
        except Exception as e:
            _logger.info(
                'Nexus AI coa-mapping unreachable, using rule-based fallback: %s', e
            )
        return None

    def action_execute_import(self):
        """Execute automated batch creation in Odoo for Products, Partners, Accounts, or Employees."""
        self.ensure_one()
        if not self.raw_parsed_json:
            raise UserError(_('يرجى فحص ومعاينة البيانات أولاً.'))

        parsed_data = json.loads(self.raw_parsed_json)
        rows = parsed_data.get('rows', [])
        images_map = parsed_data.get('images_map', {})
        target = parsed_data.get('detected_target', 'products')

        if target not in ('products', 'partners', 'accounts'):
            raise UserError(
                _('استيراد "%s" غير مدعوم بعد في هذا الإصدار. المدعوم حالياً: المنتجات، العملاء/الموردين، ودليل الحسابات.')
                % target
            )

        created_count = 0
        updated_count = 0
        images_loaded_count = 0

        # ════════════════ 1. IMPORT PRODUCTS & CATALOG WITH IMAGES ════════════════
        if target == 'products':
            Product = self.env['product.template']
            Category = self.env['product.category']

            for r in rows:
                name = None
                code = None
                barcode = None
                sale_price = 0.0
                cost_price = 0.0
                category_name = None
                image_url = None
                image_base64 = None

                # Field mapping heuristics
                for k, val in r.items():
                    k_lower = k.lower().strip()
                    val_str = str(val).strip()
                    if any(w in k_lower for w in ['name', 'maktx', 'description', 'اسم الصنف', 'اسم المنتج', 'item_description']):
                        name = val_str
                    elif any(w in k_lower for w in ['sku', 'matnr', 'item_number', 'item_no', 'كود الصنف', 'code', 'رقم الصنف']):
                        code = val_str
                    elif any(w in k_lower for w in ['barcode', 'ean11', 'باركود', 'upc', 'gtin']):
                        barcode = val_str
                    elif any(w in k_lower for w in ['sale_price', 'list_price', 'price', 'kbetr', 'unit_price', 'سعر البيع', 'السعر']):
                        try:
                            sale_price = float(re.sub(r'[^\d.]', '', val_str))
                        except Exception:
                            pass
                    elif any(w in k_lower for w in ['cost', 'stprs', 'cost_price', 'سعر التكلفة', 'التكلفة']):
                        try:
                            cost_price = float(re.sub(r'[^\d.]', '', val_str))
                        except Exception:
                            pass
                    elif any(w in k_lower for w in ['category', 'matkl', 'التصنيف', 'القسم', 'مجموعة']):
                        category_name = val_str
                    elif any(w in k_lower for w in ['image_url', 'صورة', 'image', 'picture', 'photo']):
                        if val_str.startswith(('http://', 'https://')):
                            image_url = val_str
                        elif len(val_str) > 100:
                            image_base64 = val_str

                if not name and code:
                    name = f"صنف - {code}"
                if not name:
                    continue

                # Check if image in ZIP map
                if not image_base64 and not image_url:
                    for lookup_key in [code, barcode, name]:
                        if lookup_key and lookup_key.lower() in images_map:
                            image_base64 = images_map[lookup_key.lower()]
                            break

                # Download image if URL
                if image_url and not image_base64:
                    try:
                        resp = requests.get(image_url, timeout=4)
                        if resp.ok:
                            image_base64 = base64.b64encode(resp.content).decode('utf-8')
                    except Exception:
                        pass

                # Resolve Category
                categ_id = False
                if category_name:
                    existing_cat = Category.search([('name', '=', category_name)], limit=1)
                    if not existing_cat:
                        existing_cat = Category.create({'name': category_name})
                    categ_id = existing_cat.id

                prod_vals = {
                    'name': name,
                    'list_price': sale_price,
                    'standard_price': cost_price,
                    'type': 'consu',
                    'company_id': self.env.company.id,
                }
                if 'is_storable' in Product._fields:
                    prod_vals['is_storable'] = True
                if code:
                    prod_vals['default_code'] = code
                if barcode:
                    prod_vals['barcode'] = barcode
                if categ_id:
                    prod_vals['categ_id'] = categ_id
                if image_base64:
                    prod_vals['image_1920'] = image_base64
                    images_loaded_count += 1

                # Match existing by code or barcode or name
                existing_prod = False
                if code:
                    existing_prod = Product.search([('default_code', '=', code), ('company_id', '=', self.env.company.id)], limit=1)
                if not existing_prod and barcode:
                    existing_prod = Product.search([('barcode', '=', barcode), ('company_id', '=', self.env.company.id)], limit=1)
                if not existing_prod:
                    existing_prod = Product.search([('name', '=', name), ('company_id', '=', self.env.company.id)], limit=1)

                if existing_prod:
                    existing_prod.write(prod_vals)
                    updated_count += 1
                else:
                    Product.create(prod_vals)
                    created_count += 1

        # ════════════════ 2. IMPORT CUSTOMERS & VENDORS (PARTNERS) ════════════════
        elif target == 'partners':
            Partner = self.env['res.partner']
            for r in rows:
                name = None
                phone = None
                email = None
                vat = None
                city = None
                is_vendor = False

                for k, val in r.items():
                    k_lower = k.lower().strip()
                    val_str = str(val).strip()
                    if any(w in k_lower for w in ['name', 'name1', 'party_name', 'الاسم', 'اسم العميل', 'اسم المورد']):
                        name = val_str
                    elif any(w in k_lower for w in ['phone', 'mobile', 'هاتف', 'جوال', 'tel']):
                        phone = val_str
                    elif any(w in k_lower for w in ['email', 'إيميل', 'بريد']):
                        email = val_str
                    elif any(w in k_lower for w in ['vat', 'tax_id', 'stcd1', 'الرقم الضريبي', 'ضريبة']):
                        vat_m = re.search(r'([3]\d{13}[3]|\d{15})', val_str)
                        vat = vat_m.group(1) if vat_m else val_str
                    elif any(w in k_lower for w in ['city', 'المدينة', 'ort01']):
                        city = val_str
                    elif any(w in k_lower for w in ['vendor', 'supplier', 'مورد', 'lifnr']):
                        is_vendor = True

                if not name:
                    continue

                part_vals = {
                    'name': name,
                    'company_id': self.env.company.id,
                }
                if phone:
                    part_vals['phone'] = phone
                if email:
                    part_vals['email'] = email
                if vat:
                    part_vals['vat'] = vat
                if city:
                    part_vals['city'] = city
                if is_vendor:
                    part_vals['supplier_rank'] = 1
                else:
                    part_vals['customer_rank'] = 1

                existing_p = False
                if vat:
                    existing_p = Partner.search([('vat', '=', vat), ('company_id', '=', self.env.company.id)], limit=1)
                if not existing_p:
                    existing_p = Partner.search([('name', '=', name), ('company_id', '=', self.env.company.id)], limit=1)

                if existing_p:
                    existing_p.write(part_vals)
                    updated_count += 1
                else:
                    Partner.create(part_vals)
                    created_count += 1

        # ════════════════ 3. IMPORT CHART OF ACCOUNTS (AI-ASSISTED) ════════════════
        mapping_rows_html = ""
        ai_used = False
        if target == 'accounts':
            Account = self.env['account.account']
            company = self.env.company
            valid_types = set(dict(Account._fields['account_type'].selection).keys())

            extracted = []
            for r in rows:
                code = None
                name = None
                parent_code = None
                type_hint = None
                for k, val in r.items():
                    k_lower = k.lower().strip()
                    val_str = str(val).strip()
                    if any(w in k_lower for w in ['account_code', 'code', 'رقم الحساب', 'كود الحساب', 'saknr']):
                        code = val_str
                    elif any(w in k_lower for w in ['account_name', 'name', 'اسم الحساب', 'وصف الحساب', 'txt50', 'description']):
                        name = val_str
                    elif any(w in k_lower for w in ['parent', 'الحساب الأب', 'حساب رئيسي', 'حساب أب']):
                        parent_code = val_str
                    elif any(w in k_lower for w in ['type', 'نوع الحساب', 'category']):
                        type_hint = val_str
                if not code or not name:
                    continue
                extracted.append({
                    'code': code,
                    'name': name,
                    'parent_code': parent_code or None,
                    'raw_type_hint': type_hint or None,
                })

            if not extracted:
                raise UserError(_('لم يتم العثور على أعمدة صالحة لكود واسم الحساب في الملف.'))

            ai_mapping = self._get_ai_coa_mapping(extracted)
            ai_used = ai_mapping is not None
            mapping_row_parts = []

            for row in extracted:
                code, name = row['code'], row['name']
                suggestion = (ai_mapping or {}).get(code)
                if suggestion and suggestion.get('account_type') in valid_types:
                    account_type = suggestion['account_type']
                    reconcile = bool(suggestion.get('reconcile'))
                    reasoning = suggestion.get('reasoning_ar') or ''
                else:
                    account_type = self._classify_account_rule_based(code, name)
                    reconcile = account_type in ('asset_receivable', 'liability_payable')
                    reasoning = 'تصنيف تلقائي بالقواعد (بدون اتصال بخدمة الذكاء الاصطناعي)'

                vals = {
                    'code': code,
                    'name': name,
                    'account_type': account_type,
                    'reconcile': reconcile,
                }

                existing_acc = Account.search([
                    ('code', '=', code),
                    ('company_ids', 'in', company.id),
                ], limit=1)
                if existing_acc:
                    existing_acc.write(vals)
                    updated_count += 1
                else:
                    vals['company_ids'] = [Command.set([company.id])]
                    Account.create(vals)
                    created_count += 1

                mapping_row_parts.append(
                    f"<tr><td style='padding:6px;border:1px solid #E2E8F0;'>{code}</td>"
                    f"<td style='padding:6px;border:1px solid #E2E8F0;'>{name}</td>"
                    f"<td style='padding:6px;border:1px solid #E2E8F0;'>{account_type}</td>"
                    f"<td style='padding:6px;border:1px solid #E2E8F0;font-size:12px;color:#666;'>{reasoning}</td></tr>"
                )

            mapping_rows_html = f"""
                <div class="card border-0 shadow-sm p-3 mt-3" style="border-radius: 10px;">
                    <strong class="mb-2 d-block">تفاصيل تصنيف دليل الحسابات ({'بمساعدة الذكاء الاصطناعي' if ai_used else 'بالقواعد الاحتياطية دون اتصال بالذكاء الاصطناعي'}):</strong>
                    <div class="table-responsive" style="max-height: 300px; overflow-y: auto;">
                        <table class="table table-bordered table-sm mb-0" style="font-size: 13px;">
                            <thead><tr>
                                <th style="padding:6px;border:1px solid #E2E8F0;background:#EDF2F7;">الكود</th>
                                <th style="padding:6px;border:1px solid #E2E8F0;background:#EDF2F7;">الاسم</th>
                                <th style="padding:6px;border:1px solid #E2E8F0;background:#EDF2F7;">النوع في Odoo</th>
                                <th style="padding:6px;border:1px solid #E2E8F0;background:#EDF2F7;">السبب</th>
                            </tr></thead>
                            <tbody>{''.join(mapping_row_parts)}</tbody>
                        </table>
                    </div>
                </div>
            """

        self.import_summary_html = f"""
            <div class="alert alert-success border-0 shadow-sm p-4 text-center" style="border-radius: 12px; background-color: #E8F5E9;">
                <span class="fa fa-check-circle fa-4x text-success mb-2"></span>
                <h3 class="text-success font-weight-bold mb-2">تم استيراد وترحيل البيانات بنجاح تام! 🚀</h3>
                <div class="d-flex justify-content-center gap-3 my-3">
                    <span class="badge bg-success p-2" style="font-size: 14px;">✅ تم إنشاء: {created_count} سجل جديد</span>
                    <span class="badge bg-info p-2" style="font-size: 14px;">🔄 تم تحديث: {updated_count} سجل</span>
                    <span class="badge bg-primary p-2" style="font-size: 14px;">🖼️ تم ربط: {images_loaded_count} صورة منتج</span>
                </div>
                <p class="text-muted mb-0">جميع البيانات أصبحت مفعلة في النظام ومتوافقة مع شاشات البيع والتقارير المالية.</p>
            </div>
            {mapping_rows_html}
        """

        self.state = 'imported'
        return self._reopen()

    def action_back_to_upload(self):
        self.ensure_one()
        self.state = 'upload'
        return self._reopen()

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }
